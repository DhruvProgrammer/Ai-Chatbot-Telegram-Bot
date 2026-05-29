import os
import json
import re
import asyncio
import logging
import tempfile
import base64
from datetime import datetime
from typing import Optional

import httpx
import PyPDF2
from aiogram import Bot, Dispatcher, Router, F, types
from aiogram.filters import Command, CommandStart
from aiogram.types import (
Message,
CallbackQuery,
InlineKeyboardMarkup,
InlineKeyboardButton,
BotCommand,
BotCommandScopeAllPrivateChats,
BotCommandScopeAllGroupChats,
BotCommandScopeChat,
)
from aiogram.enums import ChatType, ParseMode

# ─── Configuration ───────────────────────────────────────────────────────────

BOT_TOKEN = "5278733059:AAG0RI7zsuCfDCq1g8xb23jdtgopoeCy_LE"
OWNER_ID = 5360075159

NVIDIA_API_URL = "https://integrate.api.nvidia.com/v1/chat/completions"

PROXY_URL = ""

MODELS = {
    "nemotron": {
        "name": "Nemotron 3 Nano",
        "model": "nvidia/nemotron-3-nano-omni-30b-a3b-reasoning",
        "api_key": "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",
        "vision": True,
    },
    "deepseek": {
        "name": "DeepSeek V4 Flash",
        "model": "deepseek-ai/deepseek-v4-flash",
        "api_key": "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",
        "vision": False,
    },
    "mistral": {
        "name": "Mistral Large 675B",
        "model": "mistralai/mistral-large-3-675b-instruct-2512",
        "api_key": "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",
        "vision": False,
    },
    "llama": {
        "name": "Llama 3.3 70B",
        "model": "meta/llama-3.3-70b-instruct",
        "api_key": "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",
        "vision": False,
    },
    "qwen": {
        "name": "Qwen3 Coder 480B",
        "model": "qwen/qwen3-coder-480b-a35b-instruct",
        "api_key": "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",
        "vision": False,
    },
    "phi": {
        "name": "Phi-4 Multimodal",
        "model": "microsoft/phi-4-multimodal-instruct",
        "api_key": "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",
        "vision": True,
    },
}

DEFAULT_MODEL = "nemotron"
MAX_HISTORY = 20
DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")

SYSTEM_PROMPT = """You are a helpful AI assistant. Keep responses concise.

FORMAT: Use Telegram MarkdownV2:
- *bold* (single asterisk, NOT double **)
- _italic_ (single underscore)
- __underline__ (double underscore)
- ~strikethrough~ (single tilde)
- `inline code` (backtick)
- ```language\ncode block\n``` (triple backtick)
- [link text](URL)

RULES:
- Never use double asterisks ** for bold
- Escape special chars in text: _ * [ ] ( ) ~ ` > # + - = | { } . !
- Do NOT escape chars inside code blocks
- Keep responses short and well-structured

FILE CREATION:
[FILE: filename.ext]
content
[/FILE]

IMPORTANT:
- Never include XML tags like <system-reminder> in your response
- Never include thinking tags or meta-commentary
- Just provide the direct answer to the user's question"""

TELEGRAPH_API = "https://api.graph.org"

# MarkdownV2 escape helpers
_ESC = "\\"
_DOT = "\\."
_EXC = "\\!"
_DASH = "\\-"
_GT = "\\>"
_LPAREN = "\\("
_RPAREN = "\\)"
_HASH = "\\#"
_PLUS = "\\+"

# ─── Logging ─────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("yasir-bot")

# ─── Data Persistence ────────────────────────────────────────────────────────


def get_default_data() -> dict:
    return {
        "owner_id": OWNER_ID,
        "authorized_groups": {},
        "group_models": {},
        "conversations": {},
        "telegraph_token": None,
        "image_channel_id": None,
    }


def load_data() -> dict:
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                defaults = get_default_data()
                for key in defaults:
                    data.setdefault(key, defaults[key])
                return data
        except (json.JSONDecodeError, IOError):
            logger.warning("data.json corrupted or unreadable, creating fresh copy.")
    return get_default_data()


def save_data(data: dict):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


db = load_data()


# ─── Telegraph Integration ───────────────────────────────────────────────────


async def get_telegraph_token() -> str:
    if db.get("telegraph_token"):
        return db["telegraph_token"]
    client_kwargs = {}
    if PROXY_URL:
        client_kwargs["proxy"] = PROXY_URL
    async with httpx.AsyncClient(**client_kwargs) as client:
        resp = await client.post(
            f"{TELEGRAPH_API}/createAccount",
            data={"short_name": "AI_Bot", "author_name": "AI Bot"},
        )
        data = resp.json()
        token = data["result"]["access_token"]
        db["telegraph_token"] = token
        save_data(db)
        return token


# Channel ID for storing images
IMAGE_CHANNEL_ID = -1003707695999

def load_image_channel_id():
    """Load image channel ID from data.json."""
    global IMAGE_CHANNEL_ID
    saved = db.get("image_channel_id")
    if saved:
        IMAGE_CHANNEL_ID = saved

def save_image_channel_id(channel_id):
    """Save image channel ID to data.json."""
    global IMAGE_CHANNEL_ID
    IMAGE_CHANNEL_ID = channel_id
    db["image_channel_id"] = channel_id
    save_data(db)

async def upload_photo_to_telegraph(file_path: str) -> str:
    """Upload a photo using Telegram Bot API and return a URL."""
    
    if not IMAGE_CHANNEL_ID:
        logger.warning("IMAGE_CHANNEL_ID not set.")
        return None
    
    try:
        # Use FSInputFile for aiogram
        photo_file = types.FSInputFile(file_path)
        
        # Upload photo to Telegram channel
        result = await bot.send_photo(
            chat_id=IMAGE_CHANNEL_ID,
            photo=photo_file,
            caption="Uploaded for Telegraph"
        )
        
        # Get the file_id from the sent photo
        if result.photo:
            file_id = result.photo[-1].file_id  # Get largest size
            
            # Get file path from Telegram
            file_info = await bot.get_file(file_id)
            file_path_tg = file_info.file_path
            
            # Construct the URL
            image_url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path_tg}"
            
            logger.info(f"Photo uploaded to Telegram channel: {image_url}")
            return image_url
            
    except Exception as e:
        logger.error(f"Failed to upload photo to Telegram: {e}")
    
    return None


async def create_telegraph_page(title: str, content: str, author: str = "AI Bot", image_url: str = None, image_position: str = "above") -> str:
    token = await get_telegraph_token()

    content_nodes = []
    
    # Add image above text if specified
    if image_url and image_position == "above":
        content_nodes.append({
            "tag": "img",
            "attrs": {"src": image_url}
        })
        content_nodes.append({"tag": "p", "children": [""]})  # Empty line after image

    # Add text content
    for line in content.split("\n"):
        if line.strip():
            content_nodes.append({"tag": "p", "children": [line]})

    # Add image below text if specified
    if image_url and image_position == "below":
        content_nodes.append({"tag": "p", "children": [""]})  # Empty line before image
        content_nodes.append({
            "tag": "img",
            "attrs": {"src": image_url}
        })

    if not content_nodes:
        content_nodes.append({"tag": "p", "children": [content]})

    client_kwargs = {}
    if PROXY_URL:
        client_kwargs["proxy"] = PROXY_URL
    async with httpx.AsyncClient(**client_kwargs) as client:
        resp = await client.post(
            f"{TELEGRAPH_API}/createPage",
            data={
                "access_token": token,
                "title": title[:256],
                "content": json.dumps(content_nodes),
                "author_name": author,
            },
        )
        data = resp.json()
        return data["result"]["url"]


def extract_code_blocks(text: str) -> list[tuple[str, str]]:
    """Extract ```lang ... ``` blocks from text. Returns [(lang, code), ...]."""
    blocks = []
    pattern = r"```(\w*)\n(.*?)```"
    for match in re.finditer(pattern, text, re.DOTALL):
        lang = match.group(1) or "text"
        code = match.group(2).strip()
        blocks.append((lang, code))
    return blocks


# ─── AI API Call ─────────────────────────────────────────────────────────────


def clean_ai_response(text: str) -> str:
    if not text:
        return ""
    
    # Remove system-reminder tags (including multiline content)
    text = re.sub(r'<system-reminder>.*?</system-reminder>', '', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'</?system-reminder[^>]*>', '', text, flags=re.IGNORECASE)
    
    # Remove thinking tags
    text = re.sub(r'<thinking>.*?</thinking>', '', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'</?thinking[^>]*>', '', text, flags=re.IGNORECASE)
    
    # Remove any other XML-like tags that might be artifacts
    # But preserve HTML tags that might be intentional formatting
    text = re.sub(r'<(?!/?(?:b|strong|i|em|u|s|del|code|pre|a|br|img|p|br/)\b)[^>]+>', '', text)
    
    # Try to parse as JSON if it looks like JSON
    stripped = text.strip()
    if stripped.startswith('{') or stripped.startswith('['):
        try:
            import json
            data = json.loads(stripped)
            if isinstance(data, dict):
                # Extract text from various JSON formats
                for key in ['content', 'text', 'message', 'response', 'answer', 'result']:
                    if key in data:
                        return str(data[key]).strip()
            elif isinstance(data, list) and len(data) > 0:
                return str(data[0]).strip()
        except:
            pass
    
    # Clean up common artifacts
    text = text.strip()
    
    # Remove leading/trailing quotes if present
    if text.startswith('"') and text.endswith('"'):
        text = text[1:-1].strip()
    
    return text

async def call_ai(messages: list[dict], model_key: str, image_base64: str = None) -> str:
    model_info = MODELS.get(model_key, MODELS[DEFAULT_MODEL])

    if image_base64:
        vision_model = None
        for key, info in MODELS.items():
            if info.get("vision"):
                vision_model = info
                break
        if vision_model:
            model_info = vision_model
        else:
            return "No vision model available for image analysis."

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {model_info['api_key']}",
    }

    if image_base64:
        for msg in messages:
            if msg["role"] == "user":
                if isinstance(msg["content"], str):
                    msg["content"] = [
                        {"type": "text", "text": msg["content"]},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_base64}"}}
                    ]
                break

    payload = {
        "model": model_info["model"],
        "messages": messages,
        "max_tokens": 4096,
        "temperature": 0.7,
        "stream": False,
    }

    logger.info(f"Calling AI: model={model_info['model']}, has_image={image_base64 is not None}")

    for attempt in range(2):
        try:
            async with httpx.AsyncClient(timeout=180) as client:
                resp = await client.post(NVIDIA_API_URL, json=payload, headers=headers)
                logger.info(f"AI response status: {resp.status_code}")

                if resp.status_code == 200:
                    data = resp.json()
                    msg = data["choices"][0]["message"]
                    content = msg.get("content")
                    reasoning = msg.get("reasoning_content") or msg.get("reasoning")

                    if content:
                        return clean_ai_response(content)
                    elif reasoning:
                        return clean_ai_response(reasoning)
                    return "No response from AI."

                logger.error(f"AI API error {resp.status_code}: {resp.text[:200]}")

                if attempt < 1:
                    await asyncio.sleep(3)
                    continue

                return f"AI API error ({resp.status_code})."

        except httpx.TimeoutException:
            logger.warning(f"AI timeout, attempt {attempt + 1}/2")
            if attempt < 1:
                await asyncio.sleep(2)
                continue
            return "AI request timed out. Try again."

        except Exception as e:
            logger.error(f"AI call failed: {e}")
            return f"Error: {str(e)}"

    return "AI API error."


async def download_photo_as_base64(message: Message) -> Optional[str]:
    """Download the largest photo from a message and return as base64."""
    if not message.photo:
        return None
    photo = message.photo[-1]
    file = await bot.get_file(photo.file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
        tmp_path = tmp.name
        await bot.download_file(file.file_path, tmp_path)
    try:
        with open(tmp_path, "rb") as f:
            image_data = f.read()
        return base64.b64encode(image_data).decode("utf-8")
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


# ─── File Reading ────────────────────────────────────────────────────────────

TEXT_EXTENSIONS = {
    ".txt", ".py", ".js", ".ts", ".jsx", ".tsx", ".html", ".htm", ".css",
    ".scss", ".sass", ".less", ".json", ".xml", ".yaml", ".yml", ".toml",
    ".ini", ".cfg", ".conf", ".env", ".sh", ".bash", ".zsh", ".bat", ".cmd",
    ".ps1", ".psm1", ".md", ".markdown", ".rst", ".tex", ".log", ".csv",
    ".tsv", ".sql", ".r", ".rb", ".go", ".rs", ".java", ".kt", ".swift",
    ".c", ".cpp", ".h", ".hpp", ".cs", ".fs", ".lua", ".perl", ".pl",
    ".php", ".dart", ".scala", ".groovy", ".ex", ".exs", ".erl", ".hs",
        ".ml", ".clj", ".lisp", ".el", ".vim", ".dockerfile", ".makefile",
    ".cmake", ".gradle", ".properties", ".gitignore", ".dockerignore",
    ".htaccess", ".nginx", ".apache", ".vue", ".svelte", ".astro",
}

BINARY_EXTENSIONS = {
    ".pdf", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt",
}

ARCHIVE_EXTENSIONS = {
    ".zip", ".tar", ".gz", ".tgz", ".bz2", ".xz", ".7z",
}


async def read_text_file(file_path: str) -> str:
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except UnicodeDecodeError:
        try:
            with open(file_path, "r", encoding="latin-1") as f:
                return f.read()
        except Exception:
            return "[Could not decode file]"


async def read_pdf_file(file_path: str) -> str:
    text = ""
    try:
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        return f"[Error reading PDF: {e}]"
    return text.strip()


async def read_docx_file(file_path: str) -> str:
    try:
        import docx
        doc = docx.Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)
    except Exception as e:
        return f"[Error reading DOCX: {e}]"


async def read_xlsx_file(file_path: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        output = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            output.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    output.append(row_text)
        wb.close()
        return "\n".join(output)
    except Exception as e:
        return f"[Error reading XLSX: {e}]"


async def read_archive_file(file_path: str) -> str:
    import zipfile
    import tarfile

    output = []
    try:
        if zipfile.is_zipfile(file_path):
            with zipfile.ZipFile(file_path, "r") as zf:
                file_list = zf.namelist()
                output.append(f"ZIP archive with {len(file_list)} files:\n")
                for name in file_list[:50]:
                    output.append(f"  {name}")
                text_files = [
                    n for n in file_list
                    if os.path.splitext(n)[1].lower() in TEXT_EXTENSIONS
                    and not n.startswith("__MACOSX")
                ]
                if text_files:
                    output.append(f"\n--- Content of text files ---\n")
                    for name in text_files[:20]:
                        try:
                            with zf.open(name) as f:
                                content = f.read().decode("utf-8", errors="replace")
                            if len(content) > 2000:
                                content = content[:2000] + "\n[... truncated]"
                            output.append(f"\n>>> {name} <<<\n{content}")
                        except Exception:
                            output.append(f"\n>>> {name} <<< [unreadable]")
        elif tarfile.is_tarfile(file_path):
            with tarfile.open(file_path, "r:*") as tf:
                members = tf.getmembers()
                file_list = [m.name for m in members if m.isfile()]
                output.append(f"TAR archive with {len(file_list)} files:\n")
                for name in file_list[:50]:
                    output.append(f"  {name}")
                text_files = [
                    m for m in members
                    if m.isfile() and os.path.splitext(m.name)[1].lower() in TEXT_EXTENSIONS
                ]
                if text_files:
                    output.append(f"\n--- Content of text files ---\n")
                    for member in text_files[:20]:
                        try:
                            f = tf.extractfile(member)
                            if f:
                                content = f.read().decode("utf-8", errors="replace")
                                if len(content) > 2000:
                                    content = content[:2000] + "\n[... truncated]"
                                output.append(f"\n>>> {member.name} <<<\n{content}")
                        except Exception:
                            output.append(f"\n>>> {member.name} <<< [unreadable]")
        else:
            return "[Archive format not supported. Supported: .zip, .tar, .gz, .tgz, .bz2, .xz]"
    except Exception as e:
        return f"[Error reading archive: {e}]"

    return "\n".join(output) if output else "[Archive is empty]"


async def download_and_read_file(message: Message, document=None) -> Optional[str]:
    doc = document or (message.document if message else None)
    if not doc:
        return None

    file_name = doc.file_name or ""
    ext = os.path.splitext(file_name)[1].lower()

    all_extensions = TEXT_EXTENSIONS | BINARY_EXTENSIONS | ARCHIVE_EXTENSIONS
    if ext not in all_extensions and ext != "":
        return None

    file_size = doc.file_size or 0
    if file_size > 20 * 1024 * 1024:
        return "[File too large. Max 20MB supported.]"

    file = await bot.get_file(doc.file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp_path = tmp.name
        await bot.download_file(file.file_path, tmp_path)

    try:
        if ext in TEXT_EXTENSIONS:
            content = await read_text_file(tmp_path)
        elif ext == ".pdf":
            content = await read_pdf_file(tmp_path)
        elif ext in (".docx", ".doc"):
            content = await read_docx_file(tmp_path)
        elif ext in (".xlsx", ".xls"):
            content = await read_xlsx_file(tmp_path)
        elif ext in ARCHIVE_EXTENSIONS:
            content = await read_archive_file(tmp_path)
        else:
            content = await read_text_file(tmp_path)
    except Exception as e:
        content = f"[Error reading file: {e}]"
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    if not content:
        return "[File is empty or unreadable]"

    if len(content) > 12000:
        content = content[:12000] + "\n\n[... truncated due to length]"

    return content


# ─── Conversation History ────────────────────────────────────────────────────


def get_history(user_id: int) -> list[dict]:
    """Get conversation history for a specific user."""
    key = str(user_id)
    return db["conversations"].get(key, [])


def add_to_history(user_id: int, role: str, content: str):
    """Add to conversation history for a specific user."""
    key = str(user_id)
    if key not in db["conversations"]:
        db["conversations"][key] = []
    
    # Clean content before saving to history
    content = clean_ai_response(content)
    
    db["conversations"][key].append({"role": role, "content": content})
    if len(db["conversations"][key]) > MAX_HISTORY:
        db["conversations"][key] = db["conversations"][key][-MAX_HISTORY:]
    save_data(db)


def clear_history(user_id: int):
    """Clear conversation history for a specific user."""
    key = str(user_id)
    db["conversations"][key] = []
    save_data(db)


# ─── Bot Setup ───────────────────────────────────────────────────────────────

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()
router = Router()
dp.include_router(router)

# ─── Helpers ─────────────────────────────────────────────────────────────────


def is_owner(user_id: int) -> bool:
    return user_id == db.get("owner_id", OWNER_ID)


def is_authorized(chat_id: int) -> bool:
    return str(chat_id) in db["authorized_groups"]


def get_group_model(chat_id: int) -> str:
    model = db["group_models"].get(str(chat_id), DEFAULT_MODEL)
    if model not in MODELS:
        return DEFAULT_MODEL
    return model


def model_keyboard() -> InlineKeyboardMarkup:
    buttons = []
    for key, info in MODELS.items():
        name = info["name"]
        if info.get("vision"):
            name += " 📷"
        buttons.append(
            [InlineKeyboardButton(text=name, callback_data=f"setmodel:{key}")]
        )
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def is_bot_mentioned(message: Message) -> bool:
    if not message.entities or not message.text:
        return False
    for entity in message.entities:
        if entity.type == "mention":
            mention_text = message.text[entity.offset : entity.offset + entity.length]
            if bot.username and mention_text.lower() == f"@{bot.username.lower()}":
                return True
    return False


def strip_bot_mention(text: str) -> str:
    if bot.username:
        return re.sub(rf"@{re.escape(bot.username)}\s*", "", text, flags=re.IGNORECASE).strip()
    return text


def mentions_other_user(message: Message) -> bool:
    """Check if message mentions any user OTHER than the bot."""
    if not message.entities:
        return False
    for entity in message.entities:
        if entity.type == "text_mention" and entity.user:
            if entity.user.id != bot.id:
                return True
        elif entity.type == "mention":
            mention_text = message.text[entity.offset : entity.offset + entity.length]
            if bot.username and mention_text.lower() != f"@{bot.username.lower()}":
                return True
    return False


def escape_md(text: str) -> str:
    """Escape special characters for Telegram MarkdownV2."""
    if not text:
        return ""
    special = r"_*[]()~`>#+-=|{}.!"
    result = []
    for ch in text:
        if ch in special:
            result.append("\\")
        result.append(ch)
    return "".join(result)

# Alias for backward compatibility
md_escape = escape_md

def md_to_telegram(text: str) -> str:
    """Convert Markdown/HTML to Telegram MarkdownV2."""
    if not text:
        return ""

    # Step 1: Convert HTML to MarkdownV2
    text = re.sub(r'<br\s*/?>', '\n', text)
    text = re.sub(r'<b>(.*?)</b>', r'*\1*', text, flags=re.DOTALL)
    text = re.sub(r'<strong>(.*?)</strong>', r'*\1*', text, flags=re.DOTALL)
    text = re.sub(r'<i>(.*?)</i>', r'_\1_', text, flags=re.DOTALL)
    text = re.sub(r'<em>(.*?)</em>', r'_\1_', text, flags=re.DOTALL)
    text = re.sub(r'<u>(.*?)</u>', r'__\1__', text, flags=re.DOTALL)
    text = re.sub(r'<s>(.*?)</s>', r'~\1~', text, flags=re.DOTALL)
    text = re.sub(r'<del>(.*?)</del>', r'~\1~', text, flags=re.DOTALL)
    text = re.sub(r'<code>(.*?)</code>', r'`\1`', text, flags=re.DOTALL)
    text = re.sub(r'<pre>(.*?)</pre>', r'```\n\1\n```', text, flags=re.DOTALL)
    text = re.sub(r'<a href="(.*?)">(.*?)</a>', r'[\2](\1)', text, flags=re.DOTALL)
    text = re.sub(r'<[^>]+>', '', text)

    # Step 2: Convert standard Markdown bold to MarkdownV2
    text = re.sub(r'\*\*(.*?)\*\*', r'*\1*', text, flags=re.DOTALL)

    # Step 3: Parse and escape properly
    result = []
    i = 0
    length = len(text)
    special = set(r"_*[]()~`>#+-=|{}.!")

    while i < length:
        # Code block ```
        if i + 2 < length and text[i:i+3] == "```":
            result.append("```")
            i += 3
            end = text.find("```", i)
            if end != -1:
                code_content = text[i:end]
                escaped_code = code_content.replace("\\", "\\\\").replace("`", "\\`")
                result.append(escaped_code)
                result.append("```")
                i = end + 3
            continue

        # Inline code `
        if text[i] == "`":
            result.append("`")
            i += 1
            end = text.find("`", i)
            if end != -1:
                code_content = text[i:end]
                escaped_code = code_content.replace("\\", "\\\\").replace("`", "\\`")
                result.append(escaped_code)
                result.append("`")
                i = end + 1
            continue

        # Bold *text*
        if text[i] == "*":
            end = text.find("*", i + 1)
            if end != -1 and end > i + 1:
                inner = text[i+1:end]
                result.append("*")
                result.append(inner)
                result.append("*")
                i = end + 1
                continue

        # Italic _text_
        if text[i] == "_":
            # Check for underline __
            if i + 1 < length and text[i+1] == "_":
                end = text.find("__", i + 2)
                if end != -1 and end > i + 2:
                    inner = text[i+2:end]
                    result.append("__")
                    result.append(inner)
                    result.append("__")
                    i = end + 2
                    continue
            else:
                end = text.find("_", i + 1)
                if end != -1 and end > i + 1:
                    inner = text[i+1:end]
                    result.append("_")
                    result.append(inner)
                    result.append("_")
                    i = end + 1
                    continue

        # Strikethrough ~text~
        if text[i] == "~":
            end = text.find("~", i + 1)
            if end != -1 and end > i + 1:
                inner = text[i+1:end]
                result.append("~")
                result.append(inner)
                result.append("~")
                i = end + 1
                continue

        # Link [text](url)
        if text[i] == "[":
            bracket_end = text.find("]", i)
            if bracket_end != -1 and bracket_end + 1 < length and text[bracket_end + 1] == "(":
                paren_end = text.find(")", bracket_end + 2)
                if paren_end != -1:
                    link_text = text[i+1:bracket_end]
                    link_url = text[bracket_end+2:paren_end]
                    # Escape special chars in link text but not in URL
                    escaped_text = escape_md(link_text)
                    result.append(f"[{escaped_text}]({link_url})")
                    i = paren_end + 1
                    continue

        # Already escaped
        if text[i] == '\\' and i + 1 < length:
            result.append(text[i:i+2])
            i += 2
            continue

        # Regular character or special char to escape
        if text[i] in special:
            result.append("\\")
        result.append(text[i])
        i += 1

    return "".join(result)


def get_user_mention(user) -> str:
    name = user.first_name or user.username or "User"
    escaped_name = md_escape(name)
    return f"[{escaped_name}](tg://user?id={user.id})"


def split_text(text: str, max_len: int = 4000) -> list[str]:
    if len(text) <= max_len:
        return [text]

    parts = []
    current = ""

    paragraphs = text.split("\n\n")
    for para in paragraphs:
        if len(current) + len(para) + 2 <= max_len:
            current = f"{current}\n\n{para}" if current else para
        else:
            if current:
                parts.append(current)
            if len(para) <= max_len:
                current = para
            else:
                lines = para.split("\n")
                current = ""
                for line in lines:
                    if len(current) + len(line) + 1 <= max_len:
                        current = f"{current}\n{line}" if current else line
                    else:
                        if current:
                            parts.append(current)
                        if len(line) <= max_len:
                            current = line
                        else:
                            while len(line) > max_len:
                                parts.append(line[:max_len])
                                line = line[max_len:]
                            current = line

    if current:
        parts.append(current)

    return parts


def extract_files(text: str) -> tuple[list[tuple[str, str]], str]:
    """Extract [FILE: name]...[/FILE] blocks from AI response.
    Returns (list of (filename, content), remaining text)."""
    files = []
    pattern = r'\[FILE:\s*(.+?)\]\s*\n(.*?)\n\[/FILE\]'
    matches = list(re.finditer(pattern, text, re.DOTALL))

    for match in matches:
        filename = match.group(1).strip()
        content = match.group(2)
        content = strip_code_block(content)
        files.append((filename, content))

    remaining = re.sub(pattern, '', text, flags=re.DOTALL).strip()
    return files, remaining


def strip_code_block(text: str) -> str:
    """Remove ```lang ... ``` wrapping from code."""
    text = text.strip()
    if text.startswith("```") and text.endswith("```"):
        first_line_end = text.index("\n") if "\n" in text else len(text)
        text = text[first_line_end + 1:]
        text = text[:-3]
    return text.strip()


def create_pdf(content: str, title: str = "Document") -> bytes:
    """Create a PDF from text content."""
    from fpdf import FPDF
    
    # Clean content - remove system-reminder tags and other artifacts
    content = re.sub(r'<system-reminder>.*?</system-reminder>', '', content, flags=re.DOTALL)
    content = re.sub(r'</?system-reminder[^>]*>', '', content)
    content = re.sub(r'</?thinking[^>]*>', '', content)
    content = content.strip()
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=20)
    
    # Set margins
    pdf.set_left_margin(15)
    pdf.set_right_margin(15)
    
    # Try to set a Unicode font
    font_set = False
    try:
        pdf.add_font("DejaVu", "", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", uni=True)
        pdf.set_font("DejaVu", size=11)
        font_set = True
    except Exception:
        pass
    
    if not font_set:
        try:
            pdf.add_font("Arial", "", "C:\\Windows\\Fonts\\arial.ttf", uni=True)
            pdf.set_font("Arial", size=11)
            font_set = True
        except Exception:
            pass
    
    if not font_set:
        pdf.set_font("Helvetica", size=11)
    
    # Title
    pdf.set_font_size(16)
    # Clean title of non-ASCII chars for built-in fonts
    if not font_set:
        title = title.encode("ascii", "replace").decode()
    pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)
    
    # Content
    pdf.set_font_size(11)
    
    for line in content.split("\n"):
        if not line.strip():
            pdf.ln(3)
            continue
        try:
            if not font_set:
                line = line.encode("ascii", "replace").decode()
            pdf.multi_cell(0, 6, line)
        except Exception:
            try:
                safe_line = line.encode("ascii", "replace").decode()
                pdf.multi_cell(0, 6, safe_line)
            except:
                continue
    
    return bytes(pdf.output())


async def send_files(message: Message, files: list[tuple[str, str]]):
    """Create and send files to the user."""
    for filename, content in files:
        safe_name = re.sub(r'[^\w\-.]', '_', filename)
        ext = os.path.splitext(safe_name)[1].lower()
        tmp_path = os.path.join(tempfile.gettempdir(), f"bot_{safe_name}")

        try:
            if ext == ".pdf":
                title = os.path.splitext(safe_name)[0]
                pdf_bytes = create_pdf(content, title)
                with open(tmp_path, 'wb') as f:
                    f.write(pdf_bytes)
            else:
                with open(tmp_path, 'w', encoding='utf-8') as f:
                    f.write(content)

            await message.reply_document(
                types.FSInputFile(tmp_path, filename=safe_name),
                caption=f"Created `{md_escape(safe_name)}`",
                parse_mode=ParseMode.MARKDOWN_V2,
            )
        except Exception as e:
            await message.reply(f"Failed to create `{md_escape(safe_name)}`: {md_escape(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


async def send_ai_response(message: Message, user_text: str, model_key: str, user_id: int, image_base64: str = None):
    if image_base64:
        thinking_msg = await message.reply(f"Analyzing image{_DOT}{_DOT}{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
    else:
        thinking_msg = await message.reply(f"Thinking{_DOT}{_DOT}{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)

    history = get_history(user_id)
    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    messages.extend(history)
    messages.append({"role": "user", "content": user_text})

    # Clean user text before saving to history
    clean_user_text = clean_ai_response(user_text)
    add_to_history(user_id, "user", clean_user_text)

    response = await call_ai(messages, model_key, image_base64)
    
    # Clean response BEFORE saving to history
    response = clean_ai_response(response)
    add_to_history(user_id, "assistant", response)

    try:
        await thinking_msg.delete()
    except Exception:
        pass

    files, remaining = extract_files(response)

    if files:
        await send_files(message, files)

    if not remaining and files:
        return

    if not remaining:
        remaining = response

    response = md_to_telegram(remaining)

    mention = ""
    if message.from_user:
        mention = get_user_mention(message.from_user)

    parts = split_text(response, max_len=4000)

    for i, part in enumerate(parts):
        if i == 0 and mention:
            text = f"{mention}\n\n{part}"
        else:
            text = part

        try:
            if i == 0:
                await message.reply(text, parse_mode=ParseMode.MARKDOWN_V2)
            else:
                await message.answer(text, parse_mode=ParseMode.MARKDOWN_V2)
        except Exception:
            if i == 0:
                await message.reply(part)
            else:
                await message.answer(part)


# ─── Command Handlers ────────────────────────────────────────────────────────


@router.message(CommandStart())
async def cmd_start(message: Message):
    text = (
        f"Welcome{_EXC} I'm your *AI Group Manager Bot*{_DOT}\n\n"
        f"Add me to a group and use /authorize to activate me there{_DOT}\n"
        f"Use /settings to change the AI model{_DOT}\n"
        f"Use /q to ask me anything{_DOT}\n\n"
        f"> contact:@itznik\\_x"
    )
    await message.answer(text, parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("authorize"))
async def cmd_authorize(message: Message):
    if message.chat.type == ChatType.PRIVATE:
        await message.answer(f"This command only works in groups{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can authorize groups{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    chat_id = str(message.chat.id)
    chat_title = md_escape(message.chat.title or chat_id)

    if chat_id in db["authorized_groups"]:
        await message.answer(f"Group *{chat_title}* is already authorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    db["authorized_groups"][chat_id] = {
        "title": message.chat.title or chat_id,
        "authorized_at": datetime.now().isoformat(),
        "authorized_by": message.from_user.id,
    }
    db["group_models"][chat_id] = DEFAULT_MODEL
    save_data(db)

    model_name = md_escape(MODELS[DEFAULT_MODEL]["name"])
    await message.answer(
        f"Group *{chat_title}* has been authorized{_EXC}\n"
        f"Current AI model: *{model_name}*\n"
        f"Use /settings to change the model{_DOT}",
        parse_mode=ParseMode.MARKDOWN_V2,
    )


@router.message(Command("deauthorize", "unauthorize"))
async def cmd_deauthorize(message: Message):
    if message.chat.type == ChatType.PRIVATE:
        await message.answer(f"This command only works in groups{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can deauthorize groups{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    chat_id = str(message.chat.id)
    chat_title = md_escape(message.chat.title or chat_id)

    if chat_id not in db["authorized_groups"]:
        await message.answer(f"Group *{chat_title}* is not authorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    del db["authorized_groups"][chat_id]
    db["group_models"].pop(chat_id, None)
    db["conversations"].pop(chat_id, None)
    save_data(db)

    await message.answer(f"Group *{chat_title}* has been deauthorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("settings"))
async def cmd_settings(message: Message):
    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can change settings{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    if message.chat.type == ChatType.PRIVATE:
        await message.answer(
            f"Settings can only be changed inside a group{_DOT}\n"
            f"Go to your group and use /settings there{_DOT}",
            parse_mode=ParseMode.MARKDOWN_V2,
        )
        return

    chat_id = str(message.chat.id)
    if chat_id not in db["authorized_groups"]:
        await message.answer(f"This group is not authorized{_DOT} Use /authorize first{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    current_model = get_group_model(message.chat.id)
    current_name = md_escape(MODELS[current_model]["name"])

    await message.answer(
        f"Current model: *{current_name}*\n\nSelect a new AI model:",
        reply_markup=model_keyboard(),
        parse_mode=ParseMode.MARKDOWN_V2,
    )


@router.callback_query(F.data.startswith("setmodel:"))
async def cb_set_model(callback: CallbackQuery):
    if not is_owner(callback.from_user.id):
        await callback.answer("Only the owner can change models.", show_alert=True)
        return

    if callback.message.chat.type == ChatType.PRIVATE:
        await callback.answer("Change model inside a group.", show_alert=True)
        return

    model_key = callback.data.split(":")[1]
    if model_key not in MODELS:
        await callback.answer("Invalid model.", show_alert=True)
        return

    chat_id = str(callback.message.chat.id)
    if chat_id not in db["authorized_groups"]:
        await callback.answer("Group not authorized.", show_alert=True)
        return

    db["group_models"][chat_id] = model_key
    save_data(db)

    model_name = md_escape(MODELS[model_key]["name"])
    await callback.message.edit_text(
        f"AI model changed to *{model_name}*",
        parse_mode=ParseMode.MARKDOWN_V2,
    )
    await callback.answer(f"Model set to {MODELS[model_key]['name']}")


@router.message(Command("setchannel"))
async def cmd_setchannel(message: Message):
    """Set the channel ID for image uploads."""
    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can use this command{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return
    
    if message.chat.type != ChatType.PRIVATE:
        await message.answer(f"Use this command in DM with the bot{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return
    
    if len(message.command) < 2:
        current = IMAGE_CHANNEL_ID or "Not set"
        await message.answer(
            f"*Image Channel Settings*\n\n"
            f"Current channel: `{current}`\n\n"
            f"Usage: `/setchannel \\-1001234567890`\n\n"
            f"To get channel ID:\n"
            f"{_DASH} Forward a message from the channel to @userinfobot\n"
            f"{_DASH} Or use /id in the channel\n\n"
            f"Make sure the bot is admin in the channel{_DOT}",
            parse_mode=ParseMode.MARKDOWN_V2,
        )
        return
    
    try:
        channel_id = int(message.command[1].strip())
    except ValueError:
        await message.answer(f"Invalid channel ID{_DOT} It should be a number like \\-1001234567890{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return
    
    # Test if bot can send to the channel
    try:
        test_msg = await bot.send_message(channel_id, "Testing channel connection...")
        await test_msg.delete()
        save_image_channel_id(channel_id)
        await message.answer(f"Image channel set to `{channel_id}`{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
    except Exception as e:
        await message.answer(f"Cannot send to channel `{channel_id}`{_DOT}\nMake sure the bot is admin there{_DOT}\nError: {md_escape(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("clearhistory"))
async def cmd_clearhistory(message: Message):
    # Any user can clear their own history
    clear_history(message.from_user.id)
    await message.answer(f"Your conversation history cleared{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("clearallhistory"))
async def cmd_clearallhistory(message: Message):
    # Only owner can clear everyone's history
    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can clear all history{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return
    
    db["conversations"] = {}
    save_data(db)
    await message.answer(f"All conversation history cleared{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("model"))
async def cmd_model(message: Message):
    if message.chat.type == ChatType.PRIVATE:
        await message.answer(f"This command only works in groups{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    if not is_owner(message.from_user.id) and not is_authorized(message.chat.id):
        await message.answer(f"This group is not authorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    current = get_group_model(message.chat.id)
    name = md_escape(MODELS[current]["name"])
    await message.answer(f"Current AI model: *{name}*", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("q", "question"))
async def cmd_question(message: Message):
    if not message.text:
        return
    question_text = message.text.split(None, 1)
    if len(question_text) < 2 or not question_text[1].strip():
        await message.answer("Usage: `/q your question here`", parse_mode=ParseMode.MARKDOWN_V2)
        return

    user_text = question_text[1].strip()

    image_base64 = None

    if message.reply_to_message:
        replied = message.reply_to_message

        if replied.document:
            file_content = await download_and_read_file(message, document=replied.document)
            if file_content:
                file_name = replied.document.file_name or "file"
                user_text = f"User asked: {user_text}\n\n--- File: {file_name} ---\n{file_content}"

        elif replied.photo:
            image_base64 = await download_photo_as_base64(replied)
            user_text = f"User asked about this image: {user_text}"

        elif replied.text:
            replied_text = replied.text
            user_text = f"User asked: {user_text}\n\n--- Replied message ---\n{replied_text}"

        elif replied.caption:
            replied_text = replied.caption
            user_text = f"User asked: {user_text}\n\n--- Replied message ---\n{replied_text}"

    if message.chat.type == ChatType.PRIVATE:
        await send_ai_response(message, user_text, DEFAULT_MODEL, message.from_user.id, image_base64)
        return

    if is_owner(message.from_user.id):
        model_key = get_group_model(message.chat.id)
        await send_ai_response(message, user_text, model_key, message.from_user.id, image_base64)
        return

    if not is_authorized(message.chat.id):
        await message.answer(f"This group is not authorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    model_key = get_group_model(message.chat.id)
    await send_ai_response(message, user_text, model_key, message.from_user.id, image_base64)


@router.message(Command("post"))
async def cmd_post(message: Message):
    content = None
    title = "Shared Content"
    image_url = None
    image_position = "above"  # Default: image above text
    
    # Parse command arguments for --below flag
    command_text = message.text or ""
    if "--below" in command_text:
        image_position = "below"
        command_text = command_text.replace("--below", "").strip()
    
    # Extract content from command arguments
    if command_text:
        parts = command_text.split(None, 1)
        if len(parts) >= 2 and parts[1].strip():
            content = parts[1].strip()
            title = content[:50].split("\n")[0] or "Shared Text"
    
    # Handle reply to message
    if message.reply_to_message:
        replied = message.reply_to_message
        
        # Handle photo reply
        if replied.photo:
            try:
                # Download photo
                photo = replied.photo[-1]  # Get largest size
                file = await bot.get_file(photo.file_id)
                
                with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
                    tmp_path = tmp.name
                    await bot.download_file(file.file_path, tmp_path)
                
                try:
                    # Upload to Telegraph
                    image_url = await upload_photo_to_telegraph(tmp_path)
                finally:
                    # Clean up temp file
                    try:
                        os.unlink(tmp_path)
                    except:
                        pass
                
                # Use caption as content if available
                if not content:
                    content = replied.caption or "Shared Image"
                    title = "Shared Image"
                    
            except Exception as e:
                logger.error(f"Failed to process photo: {e}")
                await message.reply(f"Failed to process photo: {md_escape(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)
                return
        
        # Handle document reply
        elif replied.document:
            if not content:
                content = await download_and_read_file(message, document=replied.document)
                title = replied.document.file_name or "File Content"
        
        # Handle text reply
        elif replied.text:
            if not content:
                content = replied.text
                title = replied.text[:50].split("\n")[0] or "Shared Text"
        
        # Handle caption reply
        elif replied.caption:
            if not content:
                content = replied.caption
                title = "Shared Content"
    
    # Check if we have content to publish
    if not content and not image_url:
        await message.answer(
            f"Usage: reply to a message with /post or /post your text here\n"
            f"Options: --below (place image below text)\n\n"
            f"Examples:\n"
            f"/post Hello World\n"
            f"/post --below Hello World\n"
            f"Reply to photo with /post\n"
            f"Reply to photo with /post --below My caption",
            parse_mode=ParseMode.MARKDOWN_V2,
        )
        return
    
    # Default content if only image
    if not content:
        content = "Shared Image"
    
    try:
        thinking = await message.reply(f"Publishing to Telegraph{_DOT}{_DOT}{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        
        # Create Telegraph page with or without image
        url = await create_telegraph_page(
            title=title,
            content=content,
            image_url=image_url,
            image_position=image_position
        )
        
        try:
            await thinking.delete()
        except Exception:
            pass
        
        author = message.from_user.first_name if message.from_user else "User"
        escaped_author = md_escape(author)
        escaped_url = md_escape(url)
        
        # Build success message
        success_msg = f"Published to Telegraph by *{escaped_author}*{_DOT}\n\n{escaped_url}"
        if image_url:
            if image_position == "above":
                success_msg += f"\n\n_Image above text_"
            else:
                success_msg += f"\n\n_Image below text_"
        
        await message.reply(success_msg, parse_mode=ParseMode.MARKDOWN_V2)
        
    except Exception as e:
        await message.reply(f"Failed to publish: {md_escape(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("help"))
async def cmd_help(message: Message):
    text = (
        f"*AI Group Manager Bot {_DASH} Commands*\n\n"
        f"/start {_DASH} Welcome message\n"
        f"/q `question` {_DASH} Ask AI a question directly\n"
        f"/post {_DASH} Publish text, code, or images to Telegraph\n"
        f"/clearhistory {_DASH} Clear your conversation history\n"
        f"/model {_DASH} Show current AI model\n"
        f"/help {_DASH} Show this help message\n\n"
        f"*Owner Only Commands:*\n"
        f"/authorize {_DASH} Activate bot in this group\n"
        f"/deauthorize or /unauthorize {_DASH} Deactivate bot\n"
        f"/settings {_DASH} Change AI model\n"
        f"/setchannel {_DASH} Set image channel for /post\n"
        f"/clearallhistory {_DASH} Clear everyone's history\n\n"
        f"*How to use:*\n"
        f"{_DASH} Reply to my message to chat with me\n"
        f"{_DASH} Tag me with @ to ask something\n"
        f"{_DASH} Send any file to read it {_LPAREN}txt, pdf, docx, xlsx, zip, py, js, etc{_RPAREN}\n"
        f"{_DASH} Reply to any message or file with /q\n"
        f"{_DASH} Use /post to publish text or code to Telegraph\n"
        f"{_DASH} Use /post with \\_\\_below flag for image below text\n"
        f"{_DASH} Reply to photos with /post to publish them\n"
        f"{_DASH} Owner can chat directly without /q in DM"
    )
    await message.answer(text, parse_mode=ParseMode.MARKDOWN_V2)


# ─── Document Handler (File Reading) ─────────────────────────────────────────


@router.message(F.document)
async def handle_document(message: Message):
    if message.chat.type == ChatType.PRIVATE:
        file_content = await download_and_read_file(message)
        if file_content is None:
            await message.answer(f"File format not supported{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
            return

        caption = message.caption or "Summarize and explain the content of this file."
        user_text = f"{caption}\n\n--- File Content ---\n{file_content}"

        await send_ai_response(message, user_text, DEFAULT_MODEL, message.from_user.id)
        return

    if is_owner(message.from_user.id):
        file_content = await download_and_read_file(message)
        if file_content is None:
            return
        caption = message.caption or "Summarize and explain the content of this file."
        user_text = f"{caption}\n\n--- File Content ---\n{file_content}"
        model_key = get_group_model(message.chat.id)
        await send_ai_response(message, user_text, model_key, message.from_user.id)
        return

    if not is_authorized(message.chat.id):
        return

    is_reply_to_bot = (
        message.reply_to_message
        and message.reply_to_message.from_user
        and message.reply_to_message.from_user.id == bot.id
    )
    is_mentioned = is_bot_mentioned(message)

    if not is_reply_to_bot and not is_mentioned:
        return

    if is_reply_to_bot and mentions_other_user(message):
        return

    file_content = await download_and_read_file(message)
    if file_content is None:
        await message.answer(f"File format not supported{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    caption = message.caption or "Summarize and explain the content of this file."
    user_text = f"{caption}\n\n--- File Content ---\n{file_content}"

    model_key = get_group_model(message.chat.id)
    await send_ai_response(message, user_text, model_key, message.from_user.id)


@router.message(F.photo & F.caption)
async def handle_photo_with_caption(message: Message):
    caption = message.caption or ""
    if not caption.startswith("/q") and not caption.startswith("/question"):
        return

    question = caption.split(None, 1)
    user_text = question[1].strip() if len(question) > 1 else "Describe this image in detail."

    thinking_msg = await message.reply(f"Analyzing image{_DOT}{_DOT}{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)

    image_base64 = await download_photo_as_base64(message)
    if not image_base64:
        try:
            await thinking_msg.delete()
        except:
            pass
        await message.answer(f"Failed to download image{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    model_key = get_group_model(message.chat.id)
    if not is_authorized(message.chat.id) and not is_owner(message.from_user.id):
        model_key = DEFAULT_MODEL

    history = get_history(message.from_user.id)
    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    messages.extend(history)
    messages.append({"role": "user", "content": user_text})

    add_to_history(message.from_user.id, "user", f"[Image] {user_text}")

    response = await call_ai(messages, model_key, image_base64)
    
    # Clean response BEFORE saving to history
    response = clean_ai_response(response)
    add_to_history(message.from_user.id, "assistant", response)

    try:
        await thinking_msg.delete()
    except:
        pass

    response = md_to_telegram(response)
    parts = split_text(response, max_len=4000)

    for i, part in enumerate(parts):
        try:
            if i == 0:
                await message.reply(part, parse_mode=ParseMode.MARKDOWN_V2)
            else:
                await message.answer(part, parse_mode=ParseMode.MARKDOWN_V2)
        except:
            if i == 0:
                await message.reply(part)
            else:
                await message.answer(part)


@router.message(F.photo)
async def handle_photo(message: Message):
    return


# ─── Message Handler (AI Reply) ──────────────────────────────────────────────


def extract_tagged_users(message: Message) -> str:
    """Extract tagged users from message and return context string."""
    tagged_users = []
    
    if message.entities:
        for entity in message.entities:
            if entity.type == "text_mention" and entity.user:
                # User without username
                name = entity.user.first_name or "User"
                tagged_users.append(f"@{name} (ID: {entity.user.id})")
            elif entity.type == "mention":
                # User with username
                mention_text = message.text[entity.offset:entity.offset + entity.length]
                tagged_users.append(mention_text)
    
    if message.reply_to_message and message.reply_to_message.from_user:
        replied_user = message.reply_to_message.from_user
        name = replied_user.first_name or "User"
        if replied_user.username:
            tagged_users.append(f"@{replied_user.username}")
        else:
            tagged_users.append(f"@{name} (ID: {replied_user.id})")
        
        # Also include replied message text
        if message.reply_to_message.text:
            return f"[Replying to {name}: \"{message.reply_to_message.text}\"]\n"
    
    if tagged_users:
        return f"[Tagged users: {', '.join(tagged_users)}]\n"
    
    return ""


@router.message(F.text & ~F.text.startswith("/"))
async def handle_message(message: Message):
    # Owner can chat freely in private messages
    if message.chat.type == ChatType.PRIVATE and is_owner(message.from_user.id):
        user_text = message.text
        if not user_text:
            return
        model_key = DEFAULT_MODEL
        await send_ai_response(message, user_text, model_key, message.from_user.id)
        return
    
    # For groups: only respond to bot mentions or replies to bot
    is_reply_to_bot = (
        message.reply_to_message
        and message.reply_to_message.from_user
        and message.reply_to_message.from_user.id == bot.id
    )
    is_mentioned = is_bot_mentioned(message)

    if not is_reply_to_bot and not is_mentioned:
        return

    if is_reply_to_bot and mentions_other_user(message):
        return

    if not is_authorized(message.chat.id) and not is_owner(message.from_user.id):
        return

    user_text = strip_bot_mention(message.text) if is_mentioned else message.text
    if not user_text:
        return
    
    # Add context about tagged users (if any)
    tagged_context = extract_tagged_users(message)
    if tagged_context:
        user_text = tagged_context + user_text

    model_key = get_group_model(message.chat.id)
    await send_ai_response(message, user_text, model_key, message.from_user.id)


# ─── Startup ─────────────────────────────────────────────────────────────────


async def on_startup():
    await bot.delete_webhook(drop_pending_updates=True)
    me = await bot.get_me()
    bot.username = me.username
    
    # Load image channel ID
    load_image_channel_id()
    
    logger.info(f"Bot started: @{me.username} (id={me.id})")

    user_commands = [
        BotCommand(command="start", description="Start the bot"),
        BotCommand(command="q", description="Ask AI a question"),
        BotCommand(command="post", description="Publish to Telegraph"),
        BotCommand(command="clearhistory", description="Clear your conversation history"),
        BotCommand(command="help", description="Show help"),
        BotCommand(command="model", description="Show current AI model"),
    ]
    owner_commands = [
        BotCommand(command="start", description="Start the bot"),
        BotCommand(command="q", description="Ask AI a question"),
        BotCommand(command="post", description="Publish to Telegraph"),
        BotCommand(command="authorize", description="Activate bot in this group"),
        BotCommand(command="deauthorize", description="Deactivate bot in this group"),
        BotCommand(command="unauthorize", description="Deactivate bot in this group"),
        BotCommand(command="settings", description="Change AI model"),
        BotCommand(command="model", description="Show current AI model"),
        BotCommand(command="setchannel", description="Set image channel for /post"),
        BotCommand(command="clearhistory", description="Clear your conversation history"),
        BotCommand(command="clearallhistory", description="Clear everyone's history"),
        BotCommand(command="help", description="Show help"),
    ]

    await bot.set_my_commands(user_commands, scope=BotCommandScopeAllPrivateChats())
    await bot.set_my_commands(user_commands, scope=BotCommandScopeAllGroupChats())
    await bot.set_my_commands(owner_commands, scope=BotCommandScopeChat(chat_id=OWNER_ID))

    logger.info(f"Owner ID: {db.get('owner_id', OWNER_ID)}")
    logger.info(f"Authorized groups: {len(db['authorized_groups'])}")
    logger.info(f"Image channel: {IMAGE_CHANNEL_ID or 'Not set'}")
    logger.info("Bot is running!")


async def main():
    dp.startup.register(on_startup)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
