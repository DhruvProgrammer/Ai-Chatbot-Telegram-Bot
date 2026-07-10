import os
import sys
import json
import re
import asyncio
import logging
import tempfile
import base64
import html
from datetime import datetime
from typing import Optional

import httpx
import PyPDF2
from aiogram import Bot, Dispatcher, Router, F, types
from aiogram.enums import ChatType, ParseMode
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message,
    CallbackQuery,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    BotCommand,
    BotCommandScopeAllPrivateChats,
    BotCommandScopeAllGroupChats,
    BotCommandScopeChat,
    FSInputFile,
)
from aiogram.enums import ChatType, ParseMode

# ─── Configuration ───────────────────────────────────────────────────────────

BOT_TOKEN = "5278733059:AAG0RI7zsuCfDCq1g8xb23jdtgopoeCy_LE"
OWNER_ID = 5360075159

NVIDIA_API_URL = "https://integrate.api.nvidia.com/v1/chat/completions"
NVIDIA_EMBED_URL = "https://integrate.api.nvidia.com/v1/embeddings"
NVIDIA_API_KEY_ALT = "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"

PROXY_URL = ""

TAVILY_API_KEY = "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"

MODELS = {
    "nemotron": {
        "name": "Nemotron 3 Nano",
        "model": "nvidia/nemotron-3-ultra-550b-a55b",
        "api_key": NVIDIA_API_KEY_ALT,
        "vision": False,
    },
    "deepseek": {
        "name": "DeepSeek V4 Flash",
        "model": "deepseek-ai/deepseek-v4-flash",
        "api_key": NVIDIA_API_KEY_ALT,
        "vision": False,
    },
    "mistral": {
        "name": "Mistral Large 675B",
        "model": "mistralai/mistral-large-3-675b-instruct-2512",
        "api_key": NVIDIA_API_KEY_ALT,
        "vision": False,
    },
    "llama": {
        "name": "Llama 3.3 70B",
        "model": "meta/llama-3.3-70b-instruct",
        "api_key": NVIDIA_API_KEY_ALT,
        "vision": False,
    },
    "qwen": {
        "name": "Qwen3 Coder 480B",
        "model": "qwen/qwen3-coder-480b-a35b-instruct",
        "api_key": NVIDIA_API_KEY_ALT,
        "vision": False,
    },
    "phi": {
        "name": "Phi-4 Multimodal",
        "model": "microsoft/phi-4-multimodal-instruct",
        "api_key": NVIDIA_API_KEY_ALT,
        "vision": True,
    },
    "kimi": {
        "name": "Kimi K2.6",
        "model": "moonshotai/kimi-k2.6",
        "api_key": NVIDIA_API_KEY_ALT,
        "vision": False,
        "max_tokens": 16384,
        "temperature": 1.0,
        "top_p": 1.0,
    },
    "mistral-medium": {
        "name": "Mistral Medium 3.5",
        "model": "mistralai/mistral-medium-3.5-128b",
        "api_key": NVIDIA_API_KEY_ALT,
        "vision": False,
        "max_tokens": 16384,
        "temperature": 0.7,
        "top_p": 1.0,
        "extra_params": {"reasoning_effort": "high"},
    },
    "minimax": {
        "name": "MiniMax M2.7",
        "model": "minimaxai/minimax-m2.7",
        "api_key": NVIDIA_API_KEY_ALT,
        "vision": False,
        "max_tokens": 8192,
        "temperature": 1.0,
        "top_p": 0.95,
    },
    "gpt-oss": {
        "name": "GPT-OSS 120B",
        "model": "openai/gpt-oss-120b",
        "api_key": NVIDIA_API_KEY_ALT,
        "vision": False,
        "max_tokens": 4096,
        "temperature": 1.0,
        "top_p": 1.0,
        "extra_params": {"reasoning_effort": "low", "frequency_penalty": 0, "presence_penalty": 0},
    },
    "glm": {
        "name": "GLM 5.2(SABKA BAAP)",
        "model": "z-ai/glm-5.2",
        "api_key": NVIDIA_API_KEY_ALT,
        "vision": False,
        "max_tokens": 2048,
        "temperature": 0.6,
        "top_p": 1.0,
        "timeout": 240,
        "extra_params": {"chat_template_kwargs": {"enable_thinking": False}},
    },
    "google": {
        "name": "Google Gemma 4 31b",
        "model": "google/gemma-4-31b-it",
        "api_key": NVIDIA_API_KEY_ALT,
        "vision": False,
        "max_tokens": 1536,
        "temperature": 0.7,
        "top_p": 0.95,
        "no_system_role": True,
        "timeout": 240,
    },
        
}

DEFAULT_MODEL = "google"
DEFAULT_VISION_MODEL = "phi"
MAX_HISTORY = 20
DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")
AGENTS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "user_agents.json")
AGENTS_PER_PAGE = 5
MAX_AGENT_FILE_KB = 10
MAX_AGENTS = 10


class SetNameState(StatesGroup):
    waiting_alias = State()


class InstructionState(StatesGroup):
    waiting_instruction = State()


class AgentState(StatesGroup):
    waiting_file = State()
    waiting_rename = State()


SYSTEM_PROMPT = """You are a helpful AI assistant. Keep responses concise.

FORMAT: Use Telegram MarkdownV2:
- *bold* (single asterisk, NOT double **)
- _italic_ (single underscore)
- __underline__ (double underscore)
- ~strikethrough~ (single tilde)
- `inline code` (backtick)
- ```language\ncode block\n``` (triple backtick)
- [link text](URL)

RULES:
- Never use double asterisks ** for bold
- Escape special chars in text: _ * [ ] ( ) ~ ` > # + - = | { } . !
- Do NOT escape chars inside code blocks
- Keep responses short and well-structured

FILE CREATION:
[FILE: filename.ext]
content
[/FILE]

IMPORTANT:
- Never include XML tags like <system-reminder> in your response
- Never include thinking tags or meta-commentary
- Just provide the direct answer to the user's question

WEB SEARCH:
- You have a web_search tool. Use it for anything time-sensitive or current: news, prices, scores, weather, recent events, or facts you're not confident are still accurate.
- Don't use it for general knowledge, definitions, or things you already know well.
- After searching, answer directly using the results — don't just repeat them."""

TELEGRAPH_API = "https://api.graph.org"

# MarkdownV2 escape helpers
_ESC = "\\"
_DOT = "\\."
_EXC = "\\!"
_DASH = "\\-"
_GT = "\\>"
_LPAREN = "\\("
_RPAREN = "\\)"
_HASH = "\\#"
_PLUS = "\\+"

# ─── Logging ─────────────────────────────────────────────────────────────────

LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bot_logs.txt")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
# Mirror all logs to a file so /log can return everything since startup.
_log_file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
_log_file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s"))
logging.getLogger().addHandler(_log_file_handler)
logger = logging.getLogger("yasir-bot")

# ─── Data Persistence ────────────────────────────────────────────────────────


def get_default_data() -> dict:
    return {
        "owner_id": OWNER_ID,
        "authorized_groups": {},
        "authorized_users": {},  # user_id -> {"authorized_by": owner_id, "authorized_at": timestamp}
        "group_models": {},
        "global_model": DEFAULT_MODEL,
        "conversations": {},
        "telegraph_token": None,
        "image_channel_id": None,
        "bot_alias": None,
        "setname_enabled": True,
        "user_instructions": {},
        "question_mode": {},
    }


def load_data() -> dict:
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                defaults = get_default_data()
                for key in defaults:
                    data.setdefault(key, defaults[key])
                return data
        except (json.JSONDecodeError, IOError):
            logger.warning("data.json corrupted or unreadable, creating fresh copy.")
    return get_default_data()


def save_data(data: dict):
    os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


db = load_data()

SETNAME_SET = "setname:set"
SETNAME_EDIT = "setname:edit"
SETNAME_REMOVE = "setname:remove"
SETNAME_DISABLE = "setname:disable"
SETNAME_ENABLE = "setname:enable"

INSTRUCTION_SET = "instruction:set"
INSTRUCTION_REMOVE = "instruction:remove"
INSTRUCTION_SHOW = "instruction:show"

CANCEL_AI = "cancel:ai"

# ─── Agent Constants ─────────────────────────────────────────────────────────
AGENT_LIST = "agent:list"
AGENT_ADD = "agent:add"
AGENT_REMOVE = "agent:remove"
AGENT_SELECT_PREFIX = "agent:sel:"
AGENT_TOGGLE_PREFIX = "agent:toggle:"
AGENT_DONE = "agent:done"
AGENT_CANCEL = "agent:cancel"
AGENT_RENAME_DEFAULT = "agent:rename:default"
AGENT_RENAME_CUSTOM = "agent:rename:custom"
AGENT_PAGE_PREFIX = "agent:page:"
AGENT_CLOSE = "agent:close"

# ─── Question Mode Constants ─────────────────────────────────────────────────
QM_ENABLE = "qm:enable:"
QM_DISABLE = "qm:disable:"
QM_CLOSE = "qm:close:"
QM_ANSWER_PREFIX = "qm:ans:"

# ─── Instruction Guide Constant ──────────────────────────────────────────────
INSTR_CLOSE = "instr:close"

INSTRUCTION_GUIDE = """<b>🤖 Welcome to your AI Assistant</b>

This is a multi-purpose AI bot with agents, web search, translation, image understanding and more. Here's how to get started:

<b>💬 Talk to the AI</b>
• Send <code>/q your question</code> to ask anything
• Or just reply to my message, or tag me with @
• Use <code>/search</code> for live web answers with cited sources

<b>📁 Agents — /agent</b>
Agents are <code>.md</code> files (max 10KB) you upload to give the AI special instructions or context.
• <b>Add Agent</b> — send a <code>.md</code> file, then name it
• <b>Toggle</b> — tap an agent to mark it ✅ active (max 5 active). Active agents are used in your chats
• <b>Remove Agent</b> — pick agents to delete (max 5 per delete)
• <b>Close</b> — hide this menu
Your agents are private — only you can see or change them.

<b>❓ Question Mode — /questionmode</b>
When ON, the bot asks clarifying questions with inline buttons for complex or building/development tasks, and avoids guessing. <i>Owner-only setting.</i>

<b>⚙️ Other commands</b>
<code>/translate</code> · <code>/safety</code> · <code>/post</code> · <code>/model</code> · <code>/instructions</code> (set your personal AI instruction) · <code>/clearhistory</code> · <code>/help</code>

<b>🔒 Privacy</b>
Everything is per-user. One user cannot view or edit another user's agents or settings."""

QUESTION_MODE_INSTRUCTION = """QUESTION MODE is ON.
Rules:
- NEVER hallucinate. If you are unsure or lack information, say so honestly instead of guessing.
- If the user's request is complex, ambiguous, or related to building / construction / development / coding / architecture / engineering, ask clarifying questions BEFORE giving the final answer.
- You may ask several questions one at a time: after the user picks an answer, you can ask the next question, and so on, until you have enough to answer well.
- When you need to ask, output ONLY this exact format and nothing else:
[ASK] your clarifying question here [/ASK]
[OPT] first option [/OPT]
[OPT] second option [/OPT]
- Use [OPT] lines only when discrete choices help the user; keep options short.
- If no clarification is needed, just answer normally (no [ASK] tags at all)."""

# ─── Agent Data Helpers ─────────────────────────────────────────────────────

def load_agents() -> dict:
    if os.path.exists(AGENTS_FILE):
        try:
            with open(AGENTS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            logger.warning("user_agents.json corrupted, creating fresh copy.")
    return {}

def save_agents(data: dict):
    os.makedirs(os.path.dirname(AGENTS_FILE), exist_ok=True)
    with open(AGENTS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def get_user_agents(user_id: int) -> list:
    agents = load_agents()
    return agents.get(str(user_id), [])


def get_active_agent_content(user_id: int) -> str:
    """Get combined content of all active agents for a user."""
    agents = get_user_agents(user_id)
    active_content = []
    for agent in agents:
        if agent.get("active", True):
            content = agent.get("content", "").strip()
            if content:
                name = agent.get("name", agent.get("filename", "Agent"))
                active_content.append(f"--- Agent: {name} ---\n{content}")
    return "\n\n".join(active_content)


def save_user_agents(user_id: int, agents_list: list):
    all_agents = load_agents()
    all_agents[str(user_id)] = agents_list
    save_agents(all_agents)


async def _ensure_agent_private(callback: CallbackQuery) -> bool:
    """Ensure agent callbacks only run in the user's own private chat (DM with bot).

    Agents are strictly per-user. Inline buttons must never be clickable by,
    or affect, another user — so every agent interaction is confined to a
    private chat where no one else can see or press the buttons.
    """
    if callback.message is None or callback.message.chat.type != ChatType.PRIVATE:
        await callback.answer(
            "Agent settings are private and can only be used in a direct chat with the bot.",
            show_alert=True,
        )
        return False
    return True

# ─── Active Requests for Cancel Feature ──────────────────────────────────────
# Structure: {message_id: {"user_id": int, "cancelled": bool}}
active_requests = {}

# ─── Authorization Functions (Leech Bot Architecture) ──────────────────────

def _is_owner(user_id: int) -> bool:
    """Check if user is the bot owner."""
    return user_id == db.get("owner_id", OWNER_ID)

def authorize_user(user_id: int, authorized_by: int):
    """Add a user to global authorized users list."""
    db.setdefault("authorized_users", {})[str(user_id)] = {
        "authorized_by": authorized_by,
        "authorized_at": datetime.now().isoformat(),
    }
    save_data(db)

def deauthorize_user(user_id: int):
    """Remove a user from global authorized users list."""
    user_id_str = str(user_id)
    if user_id_str in db.get("authorized_users", {}):
        del db["authorized_users"][user_id_str]
        save_data(db)

def is_user_authorized(user_id: int) -> bool:
    """Check if user is globally authorized (can use bot in any chat)."""
    return str(user_id) in db.get("authorized_users", {})

def can_use_bot(user_id: int, chat_id: int) -> bool:
    """
    Check if a user can use the bot in a specific chat.
    
    Leech bot architecture:
    - Owner: Can use anywhere
    - Globally authorized users: Can use in any chat (authorized by owner with /auth @user)
    - Group members: Can only use in authorized groups
    """
    # Owner can always use
    if _is_owner(user_id):
        return True
    
    # Globally authorized users can use anywhere
    if is_user_authorized(user_id):
        return True
    
    # Check if group is authorized and user is a member
    chat_id_str = str(chat_id)
    if chat_id_str in db.get("authorized_groups", {}):
        return True
    
    return False

def is_authorized_group(chat_id: int) -> bool:
    """Check if a group is authorized."""
    return str(chat_id) in db.get("authorized_groups", {})


async def reject_unauthorized(message: Message):
    """Send a consistent rejection message to an unauthorized user."""
    await message.answer(
        f"You're not authorized to use this bot{_DOT} Ask the owner to authorize you{_DOT}",
        parse_mode=ParseMode.MARKDOWN_V2,
    )


# ─── Telegraph Integration ───────────────────────────────────────────────────


async def get_telegraph_token() -> str:
    if db.get("telegraph_token"):
        return db["telegraph_token"]
    client_kwargs = {}
    if PROXY_URL:
        client_kwargs["proxy"] = PROXY_URL
    async with httpx.AsyncClient(**client_kwargs) as client:
        resp = await client.post(
            f"{TELEGRAPH_API}/createAccount",
            data={"short_name": "AI_Bot", "author_name": "AI Bot"},
        )
        data = resp.json()
        token = data["result"]["access_token"]
        db["telegraph_token"] = token
        save_data(db)
        return token


# Channel ID for storing images
IMAGE_CHANNEL_ID = -1003707695999

def load_image_channel_id():
    """Load image channel ID from data.json."""
    global IMAGE_CHANNEL_ID
    saved = db.get("image_channel_id")
    if saved:
        IMAGE_CHANNEL_ID = saved

def save_image_channel_id(channel_id):
    """Save image channel ID to data.json."""
    global IMAGE_CHANNEL_ID
    IMAGE_CHANNEL_ID = channel_id
    db["image_channel_id"] = channel_id
    save_data(db)

async def upload_photo_to_telegraph(file_path: str) -> str:
    """Upload a photo using Telegram Bot API and return a URL."""
    
    if not IMAGE_CHANNEL_ID:
        logger.warning("IMAGE_CHANNEL_ID not set.")
        return None
    
    try:
        # Use FSInputFile for aiogram
        photo_file = types.FSInputFile(file_path)
        
        # Upload photo to Telegram channel
        result = await bot.send_photo(
            chat_id=IMAGE_CHANNEL_ID,
            photo=photo_file,
            caption="Uploaded for Telegraph"
        )
        
        # Get the file_id from the sent photo
        if result.photo:
            file_id = result.photo[-1].file_id  # Get largest size
            
            # Get file path from Telegram
            file_info = await bot.get_file(file_id)
            file_path_tg = file_info.file_path
            
            # Construct the URL
            image_url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path_tg}"
            
            logger.info(f"Photo uploaded to Telegram channel: {image_url}")
            return image_url
            
    except Exception as e:
        logger.error(f"Failed to upload photo to Telegram: {e}")
    
    return None


async def create_telegraph_page(title: str, content: str, author: str = "AI Bot", image_url: str = None, image_position: str = "above") -> str:
    token = await get_telegraph_token()

    content_nodes = []
    
    # Add image above text if specified
    if image_url and image_position == "above":
        content_nodes.append({
            "tag": "img",
            "attrs": {"src": image_url}
        })
        content_nodes.append({"tag": "p", "children": [""]})  # Empty line after image

    # Add text content
    for line in content.split("\n"):
        if line.strip():
            content_nodes.append({"tag": "p", "children": [line]})

    # Add image below text if specified
    if image_url and image_position == "below":
        content_nodes.append({"tag": "p", "children": [""]})  # Empty line before image
        content_nodes.append({
            "tag": "img",
            "attrs": {"src": image_url}
        })

    if not content_nodes:
        content_nodes.append({"tag": "p", "children": [content]})

    client_kwargs = {}
    if PROXY_URL:
        client_kwargs["proxy"] = PROXY_URL
    async with httpx.AsyncClient(**client_kwargs) as client:
        resp = await client.post(
            f"{TELEGRAPH_API}/createPage",
            data={
                "access_token": token,
                "title": title[:256],
                "content": json.dumps(content_nodes),
                "author_name": author,
            },
        )
        data = resp.json()
        return data["result"]["url"]


def extract_code_blocks(text: str) -> list[tuple[str, str]]:
    """Extract ```lang ... ``` blocks from text. Returns [(lang, code), ...]."""
    blocks = []
    pattern = r"```(\w*)\n(.*?)```"
    for match in re.finditer(pattern, text, re.DOTALL):
        lang = match.group(1) or "text"
        code = match.group(2).strip()
        blocks.append((lang, code))
    return blocks


# ─── AI API Call ─────────────────────────────────────────────────────────────


def clean_ai_response(text: str) -> str:
    if not text:
        return ""
    
    # Remove system-reminder tags (including multiline content)
    text = re.sub(r'<system-reminder>.*?</system-reminder>', '', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'</?system-reminder[^>]*>', '', text, flags=re.IGNORECASE)
    
    # Remove thinking tags
    text = re.sub(r'<thinking>.*?</thinking>', '', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'</?thinking[^>]*>', '', text, flags=re.IGNORECASE)
    
    # Remove any other XML-like tags that might be artifacts
    # But preserve HTML tags that might be intentional formatting
    text = re.sub(r'<(?!/?(?:b|strong|i|em|u|s|del|code|pre|a|br|img|p|br/)\b)[^>]+>', '', text)
    
    # Try to parse as JSON if it looks like JSON
    stripped = text.strip()
    if stripped.startswith('{') or stripped.startswith('['):
        try:
            import json
            data = json.loads(stripped)
            if isinstance(data, dict):
                # Extract text from various JSON formats
                for key in ['content', 'text', 'message', 'response', 'answer', 'result']:
                    if key in data:
                        return str(data[key]).strip()
            elif isinstance(data, list) and len(data) > 0:
                return str(data[0]).strip()
        except:
            pass
    
    # Clean up common artifacts
    text = text.strip()
    
    # Remove leading/trailing quotes if present
    if text.startswith('"') and text.endswith('"'):
        text = text[1:-1].strip()
    
    return text

WEB_SEARCH_TOOL = {
    "type": "function",
    "function": {
        "name": "web_search",
        "description": (
            "Search the live web for current, real-time, or up-to-date information: "
            "news, prices, scores, weather, recent events, current status of people "
            "or companies, or anything that may have changed since your training data "
            "or that you are not confident about. Only call this when the question "
            "genuinely needs fresh information — not for general knowledge you already know."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "A concise web search query, 2-6 words works best.",
                }
            },
            "required": ["query"],
        },
    },
}


async def _run_search_tool_calls(payload: dict, headers: dict, request_timeout: int, tool_calls: list) -> str:
    """Execute web_search tool calls requested by the model, then ask the model
    to produce a final answer using the results. Capped at one round of tool use."""
    followup_messages = list(payload["messages"])
    followup_messages.append({
        "role": "assistant",
        "content": None,
        "tool_calls": tool_calls,
    })

    for tc in tool_calls[:3]:
        func = tc.get("function", {}) or {}
        name = func.get("name", "")
        raw_args = func.get("arguments", "{}")
        try:
            args = json.loads(raw_args) if isinstance(raw_args, str) else (raw_args or {})
        except (json.JSONDecodeError, TypeError):
            args = {}
        query = (args.get("query") or "").strip()

        if name == "web_search" and query:
            logger.info(f"Model-initiated web search: {query!r}")
            search_data = await web_search(query, max_results=5)
            result_text = format_search_results_for_ai(query, search_data)
        else:
            result_text = "Error: unknown tool or missing query."

        followup_messages.append({
            "role": "tool",
            "tool_call_id": tc.get("id", ""),
            "name": name,
            "content": result_text[:4000],
        })

    followup_payload = dict(payload)
    followup_payload["messages"] = followup_messages
    followup_payload.pop("tools", None)
    followup_payload.pop("tool_choice", None)

    try:
        async with httpx.AsyncClient(timeout=request_timeout) as client:
            resp = await client.post(NVIDIA_API_URL, json=followup_payload, headers=headers)
            if resp.status_code == 200:
                data = resp.json()
                msg = data["choices"][0]["message"]
                content = msg.get("content")
                reasoning = msg.get("reasoning_content") or msg.get("reasoning")
                if content:
                    return clean_ai_response(content)
                elif reasoning:
                    return clean_ai_response(reasoning)
                return "No response from AI."
            logger.error(f"AI follow-up (post-search) error {resp.status_code}: {resp.text[:200]}")
            return f"Searched the web but the follow-up answer failed ({resp.status_code})."
    except httpx.TimeoutException:
        return "Searched the web but the follow-up answer timed out. Try again."
    except Exception as e:
        logger.error(f"AI follow-up (post-search) call failed: {e}")
        return f"Searched the web but hit an error generating the answer: {e}"


async def call_ai(messages: list[dict], model_key: str, image_base64: str = None, enable_search: bool = True) -> str:
    model_info = MODELS.get(model_key, MODELS[DEFAULT_MODEL])

    if image_base64:
        vision_model = MODELS.get(DEFAULT_VISION_MODEL)
        if not vision_model or not vision_model.get("vision"):
            # Fallback: find any model still flagged vision-capable
            vision_model = next((info for info in MODELS.values() if info.get("vision")), None)
        if vision_model:
            model_info = vision_model
        else:
            return "No vision model available for image analysis."

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {model_info['api_key']}",
    }

    if image_base64:
        for i in range(len(messages) - 1, -1, -1):
            msg = messages[i]
            if msg["role"] == "user":
                if isinstance(msg["content"], str):
                    # Replace with a new dict (don't mutate msg in place) so we
                    # never corrupt the original history object stored in db.
                    messages[i] = {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": msg["content"]},
                            {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_base64}"}}
                        ]
                    }
                break

    if model_info.get("no_system_role"):
        system_texts = []
        rest = []
        for msg in messages:
            if msg.get("role") == "system":
                c = msg.get("content")
                if isinstance(c, str):
                    system_texts.append(c)
            else:
                rest.append(msg)
        if system_texts:
            prefix = "\n\n".join(system_texts).strip()
            injected = False
            for msg in rest:
                if msg.get("role") == "user":
                    c = msg.get("content")
                    if isinstance(c, str):
                        msg["content"] = f"{prefix}\n\n{c}"
                    elif isinstance(c, list):
                        for part in c:
                            if isinstance(part, dict) and part.get("type") == "text":
                                part["text"] = f"{prefix}\n\n{part.get('text', '')}"
                                break
                        else:
                            c.insert(0, {"type": "text", "text": prefix})
                    injected = True
                    break
            if not injected:
                rest.insert(0, {"role": "user", "content": prefix})
        messages = rest

    payload = {
        "model": model_info["model"],
        "messages": messages,
        "max_tokens": model_info.get("max_tokens", 4096),
        "temperature": model_info.get("temperature", 0.7),
        "top_p": model_info.get("top_p", 1.0),
        "stream": False,
    }
    extra = {k: v for k, v in model_info.get("extra_params", {}).items()}
    payload.update(extra)

    use_search = bool(enable_search and not image_base64 and TAVILY_API_KEY)
    if use_search:
        payload["tools"] = [WEB_SEARCH_TOOL]
        payload["tool_choice"] = "auto"

    logger.info(f"Calling AI: model={model_info['model']}, has_image={image_base64 is not None}, search={use_search}")

    request_timeout = model_info.get("timeout", 180)

    for attempt in range(2):
        try:
            async with httpx.AsyncClient(timeout=request_timeout) as client:
                resp = await client.post(NVIDIA_API_URL, json=payload, headers=headers)
                logger.info(f"AI response status: {resp.status_code}")

                if resp.status_code == 200:
                    data = resp.json()
                    msg = data["choices"][0]["message"]
                    tool_calls = msg.get("tool_calls")

                    if use_search and tool_calls:
                        return await _run_search_tool_calls(payload, headers, request_timeout, tool_calls)

                    content = msg.get("content")
                    reasoning = msg.get("reasoning_content") or msg.get("reasoning")

                    if content:
                        return clean_ai_response(content)
                    elif reasoning:
                        return clean_ai_response(reasoning)
                    return "No response from AI."

                # Some models reject an unsupported "tools" param outright — drop it and retry.
                if use_search and resp.status_code in (400, 422):
                    logger.warning(f"Model {model_info['model']} rejected tools param ({resp.status_code}), retrying without search")
                    payload.pop("tools", None)
                    payload.pop("tool_choice", None)
                    use_search = False
                    continue

                logger.error(f"AI API error {resp.status_code}: {resp.text[:200]}")

                if attempt < 1:
                    await asyncio.sleep(3)
                    continue

                return f"AI API error ({resp.status_code})."

        except httpx.TimeoutException:
            logger.warning(f"AI timeout, attempt {attempt + 1}/2")
            if attempt < 1:
                await asyncio.sleep(2)
                continue
            return "AI request timed out. Try again."

        except Exception as e:
            logger.error(f"AI call failed: {e}")
            return f"Error: {str(e)}"

    return "AI API error."


async def call_translate(text: str, target_lang: str) -> str:
    """Translate text using NVIDIA riva-translate."""
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {NVIDIA_API_KEY_ALT}",
    }
    payload = {
        "model": "nvidia/riva-translate-4b-instruct-v1.1",
        "messages": [{"role": "user", "content": f"Translate English to {target_lang}: {text}"}],
        "max_tokens": 1024,
        "temperature": 0.3,
        "stream": False,
    }

    try:
        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.post(NVIDIA_API_URL, json=payload, headers=headers)
            if resp.status_code == 200:
                data = resp.json()
                return data["choices"][0]["message"]["content"].strip()
            return f"Translation API error ({resp.status_code})."
    except httpx.TimeoutException:
        return "Translation timed out. Try again."
    except Exception as e:
        return f"Translation error: {str(e)}"


async def call_safety(text: str) -> str:
    """Check content safety using NVIDIA nemotron-content-safety."""
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {NVIDIA_API_KEY_ALT}",
    }
    payload = {
        "model": "nvidia/nemotron-3.5-content-safety",
        "messages": [{"role": "user", "content": text}],
        "max_tokens": 256,
        "stream": False,
    }

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(NVIDIA_API_URL, json=payload, headers=headers)
            if resp.status_code == 200:
                data = resp.json()
                return data["choices"][0]["message"]["content"].strip()
            return f"Safety API error ({resp.status_code})."
    except httpx.TimeoutException:
        return "Safety check timed out."
    except Exception as e:
        return f"Safety error: {str(e)}"


async def download_photo_as_base64(message: Message) -> Optional[str]:
    """Download the largest photo from a message and return as base64."""
    if not message.photo:
        return None
    photo = message.photo[-1]
    file = await bot.get_file(photo.file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
        tmp_path = tmp.name
        await bot.download_file(file.file_path, tmp_path)
    try:
        with open(tmp_path, "rb") as f:
            image_data = f.read()
        return base64.b64encode(image_data).decode("utf-8")
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


# ─── Web Search (Tavily) ──────────────────────────────────────────────────────


async def web_search(query: str, max_results: int = 5) -> dict:
    """Perform a web search using Tavily and return structured results."""
    try:
        from tavily import TavilyClient

        client = TavilyClient(api_key=TAVILY_API_KEY)

        loop = asyncio.get_event_loop()
        response = await loop.run_in_executor(
            None,
            lambda: client.search(
                query=query,
                search_depth="advanced",
                max_results=max_results,
                include_answer="advanced",
                include_raw_content=False,
            ),
        )

        answer = response.get("answer") or ""
        results = response.get("results") or []

        formatted_results = []
        for r in results[:max_results]:
            formatted_results.append(
                {
                    "title": r.get("title", ""),
                    "url": r.get("url", ""),
                    "content": r.get("content", ""),
                }
            )

        return {"answer": answer, "results": formatted_results}

    except Exception as e:
        logger.error(f"Tavily search failed: {e}")
        return {"answer": "", "results": [], "error": str(e)}


def format_search_results_for_ai(query: str, search_data: dict) -> str:
    """Format Tavily search results into a context block for the AI."""
    parts = [f"[Web search performed for: \"{query}\"]\n"]

    if search_data.get("answer"):
        parts.append(f"AI-generated summary from web:\n{search_data['answer']}\n")

    results = search_data.get("results") or []
    if results:
        parts.append("Sources:\n")
        for i, r in enumerate(results, 1):
            parts.append(f"{i}. {r['title']}\n   URL: {r['url']}\n   {r['content']}\n")

    if not search_data.get("answer") and not results:
        parts.append("No results found on the web.\n")

    return "\n".join(parts)


def _html_escape(text: str) -> str:
    if not text:
        return ""
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _extract_domain(url: str) -> str:
    if not url:
        return ""
    try:
        from urllib.parse import urlparse
        netloc = urlparse(url).netloc or ""
        if netloc.startswith("www."):
            netloc = netloc[4:]
        return netloc
    except Exception:
        return ""


def format_search_sources_html(query: str, search_data: dict, ai_model_name: str = "") -> str:
    """Build a Telegram HTML block for /search sources. Safe for any URL/title."""
    results = search_data.get("results") or []
    if not results:
        return ""

    lines = []
    q_display = query.strip()
    if len(q_display) > 120:
        q_display = q_display[:117].rstrip() + "..."
    lines.append(f"🔎 <b>Query:</b> <i>{_html_escape(q_display)}</i>")
    lines.append("━━━━━━━━━━━━━━━━━")
    lines.append("📚 <b>Sources</b>")
    lines.append("")

    for i, r in enumerate(results[:5], 1):
        title = re.sub(r"\s+", " ", (r.get("title") or "").strip()) or "Source"
        if len(title) > 110:
            title = title[:107].rstrip() + "..."
        url = (r.get("url") or "").strip()
        domain = _extract_domain(url)
        content = re.sub(r"\s+", " ", (r.get("content") or "").strip())
        if len(content) > 180:
            content = content[:177].rstrip() + "..."
        score = r.get("score")

        header = f"<b>{i}.</b> "
        if url:
            header += f"<a href=\"{_html_escape(url)}\">{_html_escape(title)}</a>"
        else:
            header += f"<b>{_html_escape(title)}</b>"
        lines.append(header)

        meta_bits = []
        if domain:
            meta_bits.append(f"<code>{_html_escape(domain)}</code>")
        if isinstance(score, (int, float)):
            pct = max(0, min(100, int(round(score * 100))))
            meta_bits.append(f"⭐ {pct}%")
        if meta_bits:
            lines.append("   " + " • ".join(meta_bits))
        if content:
            lines.append(f"   <i>{_html_escape(content)}</i>")
        lines.append("")

    footer_bits = []
    rt = search_data.get("response_time")
    if isinstance(rt, (int, float)):
        footer_bits.append(f"⏱ {rt:.2f}s")
    footer_bits.append(f"📄 {len(results)} results")
    if ai_model_name:
        footer_bits.append(f"🤖 {_html_escape(ai_model_name)}")
    lines.append("━━━━━━━━━━━━━━━━━")
    lines.append(" • ".join(footer_bits))

    return "\n".join(lines).rstrip()


def format_search_results_for_user(search_data: dict, ai_answer: str = None) -> str:
    """Format search results for direct display to the user (without AI)."""
    parts = []

    if ai_answer:
        parts.append(f"*Answer:*\n{ai_answer}\n")

    answer = search_data.get("answer")
    if answer and answer != ai_answer:
        parts.append(f"*Summary:*\n{answer}\n")

    results = search_data.get("results") or []
    if results:
        parts.append("*Sources:*\n")
        for i, r in enumerate(results, 1):
            title = r["title"].replace("`", "'").replace("[", "(").replace("]", ")")
            url = r["url"]
            parts.append(f"{i}. [{title}]({url})")

    return "\n\n".join(parts) if parts else "No results found."


# ─── File Reading ────────────────────────────────────────────────────────────

TEXT_EXTENSIONS = {
    ".txt", ".py", ".js", ".ts", ".jsx", ".tsx", ".html", ".htm", ".css",
    ".scss", ".sass", ".less", ".json", ".xml", ".yaml", ".yml", ".toml",
    ".ini", ".cfg", ".conf", ".env", ".sh", ".bash", ".zsh", ".bat", ".cmd",
    ".ps1", ".psm1", ".md", ".markdown", ".rst", ".tex", ".log", ".csv",
    ".tsv", ".sql", ".r", ".rb", ".go", ".rs", ".java", ".kt", ".swift",
    ".c", ".cpp", ".h", ".hpp", ".cs", ".fs", ".lua", ".perl", ".pl",
    ".php", ".dart", ".scala", ".groovy", ".ex", ".exs", ".erl", ".hs",
        ".ml", ".clj", ".lisp", ".el", ".vim", ".dockerfile", ".makefile",
    ".cmake", ".gradle", ".properties", ".gitignore", ".dockerignore",
    ".htaccess", ".nginx", ".apache", ".vue", ".svelte", ".astro",
}

BINARY_EXTENSIONS = {
    ".pdf", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt",
}

ARCHIVE_EXTENSIONS = {
    ".zip", ".tar", ".gz", ".tgz", ".bz2", ".xz", ".7z",
}


async def read_text_file(file_path: str) -> str:
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except UnicodeDecodeError:
        try:
            with open(file_path, "r", encoding="latin-1") as f:
                return f.read()
        except Exception:
            return "[Could not decode file]"


async def read_pdf_file(file_path: str) -> str:
    text = ""
    try:
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        return f"[Error reading PDF: {e}]"
    return text.strip()


async def read_docx_file(file_path: str) -> str:
    try:
        import docx
        doc = docx.Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)
    except Exception as e:
        return f"[Error reading DOCX: {e}]"


async def read_xlsx_file(file_path: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        output = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            output.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    output.append(row_text)
        wb.close()
        return "\n".join(output)
    except Exception as e:
        return f"[Error reading XLSX: {e}]"


async def read_archive_file(file_path: str) -> str:
    import zipfile
    import tarfile

    output = []
    try:
        if zipfile.is_zipfile(file_path):
            with zipfile.ZipFile(file_path, "r") as zf:
                file_list = zf.namelist()
                output.append(f"ZIP archive with {len(file_list)} files:\n")
                for name in file_list[:50]:
                    output.append(f"  {name}")
                text_files = [
                    n for n in file_list
                    if os.path.splitext(n)[1].lower() in TEXT_EXTENSIONS
                    and not n.startswith("__MACOSX")
                ]
                if text_files:
                    output.append(f"\n--- Content of text files ---\n")
                    for name in text_files[:20]:
                        try:
                            with zf.open(name) as f:
                                content = f.read().decode("utf-8", errors="replace")
                            if len(content) > 2000:
                                content = content[:2000] + "\n[... truncated]"
                            output.append(f"\n>>> {name} <<<\n{content}")
                        except Exception:
                            output.append(f"\n>>> {name} <<< [unreadable]")
        elif tarfile.is_tarfile(file_path):
            with tarfile.open(file_path, "r:*") as tf:
                members = tf.getmembers()
                file_list = [m.name for m in members if m.isfile()]
                output.append(f"TAR archive with {len(file_list)} files:\n")
                for name in file_list[:50]:
                    output.append(f"  {name}")
                text_files = [
                    m for m in members
                    if m.isfile() and os.path.splitext(m.name)[1].lower() in TEXT_EXTENSIONS
                ]
                if text_files:
                    output.append(f"\n--- Content of text files ---\n")
                    for member in text_files[:20]:
                        try:
                            f = tf.extractfile(member)
                            if f:
                                content = f.read().decode("utf-8", errors="replace")
                                if len(content) > 2000:
                                    content = content[:2000] + "\n[... truncated]"
                                output.append(f"\n>>> {member.name} <<<\n{content}")
                        except Exception:
                            output.append(f"\n>>> {member.name} <<< [unreadable]")
        else:
            return "[Archive format not supported. Supported: .zip, .tar, .gz, .tgz, .bz2, .xz]"
    except Exception as e:
        return f"[Error reading archive: {e}]"

    return "\n".join(output) if output else "[Archive is empty]"


async def download_and_read_file(message: Message, document=None) -> Optional[str]:
    doc = document or (message.document if message else None)
    if not doc:
        return None

    file_name = doc.file_name or ""
    ext = os.path.splitext(file_name)[1].lower()

    all_extensions = TEXT_EXTENSIONS | BINARY_EXTENSIONS | ARCHIVE_EXTENSIONS
    if ext not in all_extensions and ext != "":
        return None

    file_size = doc.file_size or 0
    if file_size > 20 * 1024 * 1024:
        return "[File too large. Max 20MB supported.]"

    file = await bot.get_file(doc.file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp_path = tmp.name
        await bot.download_file(file.file_path, tmp_path)

    try:
        if ext in TEXT_EXTENSIONS:
            content = await read_text_file(tmp_path)
        elif ext == ".pdf":
            content = await read_pdf_file(tmp_path)
        elif ext in (".docx", ".doc"):
            content = await read_docx_file(tmp_path)
        elif ext in (".xlsx", ".xls"):
            content = await read_xlsx_file(tmp_path)
        elif ext in ARCHIVE_EXTENSIONS:
            content = await read_archive_file(tmp_path)
        else:
            content = await read_text_file(tmp_path)
    except Exception as e:
        content = f"[Error reading file: {e}]"
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    if not content:
        return "[File is empty or unreadable]"

    if len(content) > 12000:
        content = content[:12000] + "\n\n[... truncated due to length]"

    return content


# ─── Conversation History ────────────────────────────────────────────────────


def get_history(user_id: int) -> list[dict]:
    """Get conversation history for a specific user."""
    key = str(user_id)
    return db["conversations"].get(key, [])


def add_to_history(user_id: int, role: str, content: str):
    """Add to conversation history for a specific user."""
    key = str(user_id)
    if key not in db["conversations"]:
        db["conversations"][key] = []
    
    # Clean content before saving to history
    content = clean_ai_response(content)
    
    db["conversations"][key].append({"role": role, "content": content})
    if len(db["conversations"][key]) > MAX_HISTORY:
        db["conversations"][key] = db["conversations"][key][-MAX_HISTORY:]
    save_data(db)


def clear_history(user_id: int):
    """Clear conversation history for a specific user."""
    key = str(user_id)
    db["conversations"][key] = []
    save_data(db)


# ─── Bot Setup ───────────────────────────────────────────────────────────────

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())
router = Router()
dp.include_router(router)

# ─── Helpers ─────────────────────────────────────────────────────────────────


def is_owner(user_id: int) -> bool:
    return user_id == db.get("owner_id", OWNER_ID)


def get_saved_bot_alias() -> Optional[str]:
    alias = db.get("bot_alias")
    return alias.strip() if isinstance(alias, str) and alias.strip() else None


def is_setname_enabled() -> bool:
    return db.get("setname_enabled", True) is not False


def get_bot_alias() -> Optional[str]:
    if not is_setname_enabled():
        return None
    return get_saved_bot_alias()


def normalize_bot_alias(name: str) -> Optional[str]:
    normalized = re.sub(r"\s+", " ", name or "").strip()
    if not normalized:
        return None
    return normalized[:40]


def _levenshtein_distance(a: str, b: str) -> int:
    if abs(len(a) - len(b)) > 1:
        return 2

    previous = list(range(len(b) + 1))
    for i, char_a in enumerate(a, start=1):
        current = [i]
        for j, char_b in enumerate(b, start=1):
            current.append(min(
                previous[j] + 1,
                current[j - 1] + 1,
                previous[j - 1] + (char_a != char_b),
            ))
        previous = current
    return previous[-1]


def _find_alias_spans(text: str, alias: str) -> list[tuple[int, int]]:
    alias = alias.strip()
    if not alias:
        return []

    exact_pattern = rf"(?<!\w){re.escape(alias)}(?!\w)"
    spans = [(match.start(), match.end()) for match in re.finditer(exact_pattern, text, flags=re.IGNORECASE)]
    if spans or " " in alias or any(not char.isalnum() for char in alias):
        return spans

    alias_lower = alias.lower()
    for match in re.finditer(r"\b\w+\b", text):
        token = match.group(0)
        if len(alias_lower) < 4 or abs(len(token) - len(alias_lower)) > 1:
            continue
        if _levenshtein_distance(token.lower(), alias_lower) <= 1:
            spans.append((match.start(), match.end()))

    return spans


def strip_bot_alias(text: str) -> Optional[str]:
    alias = get_bot_alias()
    if not alias or not text:
        return None

    spans = _find_alias_spans(text, alias)
    if not spans:
        return None

    cleaned = text
    for start, end in sorted(spans, reverse=True):
        cleaned = f"{cleaned[:start]} {cleaned[end:]}"

    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    cleaned = re.sub(r"^\s*[,.;:!?-]+\s*", "", cleaned)
    return cleaned or None


def is_alias_message(message: Message) -> bool:
    return bool(message.text and strip_bot_alias(message.text))


def get_user_instruction(user_id: int) -> Optional[str]:
    instruction = db.get("user_instructions", {}).get(str(user_id))
    return instruction.strip() if isinstance(instruction, str) and instruction.strip() else None


def set_user_instruction(user_id: int, instruction: str):
    db.setdefault("user_instructions", {})[str(user_id)] = instruction.strip()
    save_data(db)


def remove_user_instruction(user_id: int):
    db.setdefault("user_instructions", {}).pop(str(user_id), None)
    save_data(db)


# ─── Question Mode Helpers ───────────────────────────────────────────────────

def is_question_mode(user_id: int) -> bool:
    """Return whether Question Mode is enabled for a specific user."""
    qm = db.get("question_mode", {})
    if isinstance(qm, bool):
        return qm
    return bool(qm.get(str(user_id), False))


def set_question_mode(user_id: int, enabled: bool):
    qm = db.get("question_mode", {})
    if isinstance(qm, bool):
        qm = {}
    qm[str(user_id)] = bool(enabled)
    db["question_mode"] = qm
    save_data(db)


def parse_ask_block(text: str):
    """Extract a [ASK]...[/ASK] clarifying question and [OPT] options from AI text.

    Returns (question, options_list, preamble_before_ask). question is None if
    no [ASK] block is present.
    """
    m = re.search(r"\[ASK\](.*?)\[/ASK\]", text, re.DOTALL | re.IGNORECASE)
    if not m:
        return None, [], text
    question = m.group(1).strip()
    preamble = text[:m.start()].strip()
    opts = re.findall(r"\[OPT\](.*?)\[/OPT\]", text, re.DOTALL | re.IGNORECASE)
    options = [o.strip() for o in opts if o.strip()]
    return question, options, preamble


# Pending Question-Mode follow-ups: user_id -> {"chat_id": int, "options": [str]}
pending_questions = {}


def instruction_keyboard() -> InlineKeyboardMarkup:
    buttons = [
        [
            InlineKeyboardButton(text="Set instruction", callback_data=INSTRUCTION_SET),
            InlineKeyboardButton(text="Remove instruction", callback_data=INSTRUCTION_REMOVE),
        ],
        [InlineKeyboardButton(text="Show current instruction", callback_data=INSTRUCTION_SHOW)],
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def is_authorized(chat_id: int, user_id: int = None) -> bool:
    """Check if a user can use the bot in a chat."""
    if user_id is not None:
        return can_use_bot(user_id, chat_id)
    return is_authorized_group(chat_id)


def get_group_model(chat_id: int) -> str:
    model = db.get("global_model", DEFAULT_MODEL)
    if model not in MODELS:
        return DEFAULT_MODEL
    return model


MODELS_PER_PAGE = 5


def model_keyboard(page: int = 0) -> InlineKeyboardMarkup:
    items = list(MODELS.items())
    total = len(items)
    total_pages = max(1, (total + MODELS_PER_PAGE - 1) // MODELS_PER_PAGE)
    if page < 0:
        page = 0
    if page >= total_pages:
        page = total_pages - 1

    start = page * MODELS_PER_PAGE
    end = start + MODELS_PER_PAGE
    page_items = items[start:end]

    buttons = []
    for key, info in page_items:
        name = info["name"]
        if info.get("vision"):
            name += " 📷"
        buttons.append(
            [InlineKeyboardButton(text=name, callback_data=f"setmodel:{key}")]
        )

    nav_row = []
    if page > 0:
        nav_row.append(
            InlineKeyboardButton(text="⬅️ Prev", callback_data=f"modelpage:{page - 1}")
        )
    nav_row.append(
        InlineKeyboardButton(text=f"Page {page + 1}/{total_pages}", callback_data="modelpage:noop")
    )
    if page < total_pages - 1:
        nav_row.append(
            InlineKeyboardButton(text="Next ➡️", callback_data=f"modelpage:{page + 1}")
        )
    buttons.append(nav_row)

    return InlineKeyboardMarkup(inline_keyboard=buttons)


def alias_keyboard() -> InlineKeyboardMarkup:
    if not is_setname_enabled():
        buttons = [
            [InlineKeyboardButton(text="Enable /setname", callback_data=SETNAME_ENABLE)],
        ]
        return InlineKeyboardMarkup(inline_keyboard=buttons)

    buttons = [
        [
            InlineKeyboardButton(text="Set name", callback_data=SETNAME_SET),
            InlineKeyboardButton(text="Edit name", callback_data=SETNAME_EDIT),
        ],
        [
            InlineKeyboardButton(text="Remove name", callback_data=SETNAME_REMOVE),
            InlineKeyboardButton(text="Disable /setname", callback_data=SETNAME_DISABLE),
        ],
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def is_bot_mentioned(message: Message) -> bool:
    if not message.entities or not message.text:
        return False
    for entity in message.entities:
        if entity.type == "mention":
            mention_text = message.text[entity.offset : entity.offset + entity.length]
            if bot.username and mention_text.lower() == f"@{bot.username.lower()}":
                return True
    return False


def strip_bot_mention(text: str) -> str:
    if bot.username:
        return re.sub(rf"@{re.escape(bot.username)}\s*", "", text, flags=re.IGNORECASE).strip()
    return text


def mentions_other_user(message: Message) -> bool:
    """Check if message mentions any user OTHER than the bot."""
    if not message.entities:
        return False
    for entity in message.entities:
        if entity.type == "text_mention" and entity.user:
            if entity.user.id != bot.id:
                return True
        elif entity.type == "mention":
            mention_text = message.text[entity.offset : entity.offset + entity.length]
            if bot.username and mention_text.lower() != f"@{bot.username.lower()}":
                return True
    return False


def escape_md(text: str) -> str:
    """Escape special characters for Telegram MarkdownV2."""
    if not text:
        return ""
    special = r"_*[]()~`>#+-=|{}.!"
    result = []
    for ch in text:
        if ch in special:
            result.append("\\")
        result.append(ch)
    return "".join(result)


def escape_html(text: str) -> str:
    """Escape special characters for Telegram HTML."""
    if not text:
        return ""
    return text.replace("&", "&").replace("<", "<").replace(">", ">").replace('"', """).replace("'", "'")


# Alias for backward compatibility
md_escape = escape_md

def md_to_telegram(text: str) -> str:
    """Convert Markdown/HTML to Telegram MarkdownV2."""
    if not text:
        return ""

    # Step 1: Convert HTML to MarkdownV2
    text = re.sub(r'<br\s*/?>', '\n', text)
    text = re.sub(r'<b>(.*?)</b>', r'*\1*', text, flags=re.DOTALL)
    text = re.sub(r'<strong>(.*?)</strong>', r'*\1*', text, flags=re.DOTALL)
    text = re.sub(r'<i>(.*?)</i>', r'_\1_', text, flags=re.DOTALL)
    text = re.sub(r'<em>(.*?)</em>', r'_\1_', text, flags=re.DOTALL)
    text = re.sub(r'<u>(.*?)</u>', r'__\1__', text, flags=re.DOTALL)
    text = re.sub(r'<s>(.*?)</s>', r'~\1~', text, flags=re.DOTALL)
    text = re.sub(r'<del>(.*?)</del>', r'~\1~', text, flags=re.DOTALL)
    text = re.sub(r'<code>(.*?)</code>', r'`\1`', text, flags=re.DOTALL)
    text = re.sub(r'<pre>(.*?)</pre>', r'```\n\1\n```', text, flags=re.DOTALL)
    text = re.sub(r'<a href="(.*?)">(.*?)</a>', r'[\2](\1)', text, flags=re.DOTALL)
    text = re.sub(r'<[^>]+>', '', text)

    # Step 2: Convert standard Markdown bold to MarkdownV2
    text = re.sub(r'\*\*(.*?)\*\*', r'*\1*', text, flags=re.DOTALL)

    # Step 2b: Convert markdown headings (###, ##, #) into bold lines
    def _heading_to_bold(match: "re.Match") -> str:
        prefix = match.group(1)
        content = match.group(2).strip()
        return f"{prefix}*{content}*"

    text = re.sub(r'(^|\n)#{1,6}\s+([^\n]+)', _heading_to_bold, text)

    # Step 2c: Normalize markdown bullets (-, *, +) at line starts to a bullet char
    text = re.sub(r'(^|\n)[ \t]*[\-\*\+][ \t]+', r'\1• ', text)

    # Step 3: Parse and escape properly
    result = []
    i = 0
    length = len(text)
    special = set(r"_*[]()~`>#+-=|{}.!")

    while i < length:
        # Code block ```
        if i + 2 < length and text[i:i+3] == "```":
            result.append("```")
            i += 3
            end = text.find("```", i)
            if end != -1:
                code_content = text[i:end]
                escaped_code = code_content.replace("\\", "\\\\").replace("`", "\\`")
                result.append(escaped_code)
                result.append("```")
                i = end + 3
            continue

        # Inline code `
        if text[i] == "`":
            result.append("`")
            i += 1
            end = text.find("`", i)
            if end != -1:
                code_content = text[i:end]
                escaped_code = code_content.replace("\\", "\\\\").replace("`", "\\`")
                result.append(escaped_code)
                result.append("`")
                i = end + 1
            continue

        # Bold *text*
        if text[i] == "*":
            end = text.find("*", i + 1)
            if end != -1 and end > i + 1 and "\n" not in text[i+1:end]:
                inner = text[i+1:end]
                result.append("*")
                result.append(escape_md(inner))
                result.append("*")
                i = end + 1
                continue

        # Italic _text_
        if text[i] == "_":
            # Check for underline __
            if i + 1 < length and text[i+1] == "_":
                end = text.find("__", i + 2)
                if end != -1 and end > i + 2 and "\n" not in text[i+2:end]:
                    inner = text[i+2:end]
                    result.append("__")
                    result.append(escape_md(inner))
                    result.append("__")
                    i = end + 2
                    continue
            else:
                end = text.find("_", i + 1)
                if end != -1 and end > i + 1 and "\n" not in text[i+1:end]:
                    inner = text[i+1:end]
                    result.append("_")
                    result.append(escape_md(inner))
                    result.append("_")
                    i = end + 1
                    continue

        # Strikethrough ~text~
        if text[i] == "~":
            end = text.find("~", i + 1)
            if end != -1 and end > i + 1 and "\n" not in text[i+1:end]:
                inner = text[i+1:end]
                result.append("~")
                result.append(escape_md(inner))
                result.append("~")
                i = end + 1
                continue

        # Link [text](url)
        if text[i] == "[":
            bracket_end = text.find("]", i)
            if bracket_end != -1 and bracket_end + 1 < length and text[bracket_end + 1] == "(":
                paren_end = text.find(")", bracket_end + 2)
                if paren_end != -1:
                    link_text = text[i+1:bracket_end]
                    link_url = text[bracket_end+2:paren_end]
                    # Escape special chars in link text but not in URL
                    escaped_text = escape_md(link_text)
                    result.append(f"[{escaped_text}]({link_url})")
                    i = paren_end + 1
                    continue

        # Already escaped
        if text[i] == '\\' and i + 1 < length:
            result.append(text[i:i+2])
            i += 2
            continue

        # Regular character or special char to escape
        if text[i] in special:
            result.append("\\")
        result.append(text[i])
        i += 1

    return "".join(result)


# ─── Per-Model Markdown Formatters ──────────────────────────────────────────────

def _convert_italic(text: str) -> str:
    """Convert *italic* (single asterisk) to _italic_ (MarkdownV2). Must run BEFORE bold conversion."""
    return re.sub(r'(?<!\*)\*(?!\*)([^*\n]+?)\*(?!\*)', r'_\1_', text)


def _convert_bold(text: str) -> str:
    """Convert **bold** (double asterisk) to *bold* (MarkdownV2)."""
    return re.sub(r'\*\*(.*?)\*\*', r'*\1*', text, flags=re.DOTALL)


def _convert_headings(text: str) -> str:
    """Convert markdown headings (#, ##, ###) to bold lines."""
    return re.sub(r'(^|\n)#{1,6}\s+([^\n]+)', r'\1*\2*', text)


def _convert_bullets(text: str) -> str:
    """Normalize markdown bullets (-, *, +) at line starts to a bullet char."""
    return re.sub(r'(^|\n)[ \t]*[\-\*\+][ \t]+', r'\1• ', text)


def format_for_gemma(text: str) -> str:
    """Format for Google Gemma - uses **bold**, _italic_, `code`, ```code blocks```"""
    if not text:
        return ""
    # Gemma uses _italic_ already (compatible). Handle *italic* just in case.
    text = _convert_italic(text)
    # Gemma uses **bold** which needs conversion to *bold*
    text = _convert_bold(text)
    text = _convert_headings(text)
    text = _convert_bullets(text)
    return md_to_telegram(text)


def format_for_deepseek(text: str) -> str:
    """Format for DeepSeek V3/V4 - uses **bold**, *italic*, `code`, ```code```, citations [citation:X]"""
    if not text:
        return ""
    # DeepSeek uses *italic* - convert to _italic_ FIRST (before bold)
    text = _convert_italic(text)
    # DeepSeek uses **bold**
    text = _convert_bold(text)
    # Handle citations [citation:X] - keep as-is
    text = re.sub(r'\[citation:(\d+)\]', r'[citation:\1]', text)
    text = _convert_headings(text)
    text = _convert_bullets(text)
    return md_to_telegram(text)


def format_for_kimi(text: str) -> str:
    """Format for Kimi K2 - uses **bold**, *italic*, `code`, ```code```, JSON mode output"""
    if not text:
        return ""
    text = _convert_italic(text)
    text = _convert_bold(text)
    # Kimi may output JSON in text mode - detect and format
    stripped = text.strip()
    if stripped.startswith('{') and stripped.endswith('}'):
        try:
            import json
            data = json.loads(stripped)
            text = "```json\n" + json.dumps(data, indent=2, ensure_ascii=False) + "\n```"
        except Exception:
            pass
    text = _convert_headings(text)
    text = _convert_bullets(text)
    return md_to_telegram(text)


def format_for_glm(text: str) -> str:
    """Format for GLM 5.2 - uses **bold**, *italic_, `code`, ```code```, thinking blocks"""
    if not text:
        return ""
    text = _convert_italic(text)
    text = _convert_bold(text)
    # Remove thinking/reasoning tags if present
    text = re.sub(r'<thinking>.*?</thinking>', '', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<reasoning>.*?</reasoning>', '', text, flags=re.DOTALL | re.IGNORECASE)
    text = _convert_headings(text)
    text = _convert_bullets(text)
    return md_to_telegram(text)


def format_for_nemotron(text: str) -> str:
    """Format for Nemotron 3 Ultra - uses **bold**, *italic*, `code`, ```code```"""
    if not text:
        return ""
    text = _convert_italic(text)
    text = _convert_bold(text)
    text = _convert_headings(text)
    text = _convert_bullets(text)
    return md_to_telegram(text)


def format_for_mistral(text: str) -> str:
    """Format for Mistral Large - uses **bold**, *italic*, `code`, ```code```"""
    if not text:
        return ""
    text = _convert_italic(text)
    text = _convert_bold(text)
    text = _convert_headings(text)
    text = _convert_bullets(text)
    return md_to_telegram(text)


def format_for_llama(text: str) -> str:
    """Format for Llama 3.3 - uses **bold**, *italic*, `code`, ```code```"""
    if not text:
        return ""
    text = _convert_italic(text)
    text = _convert_bold(text)
    text = _convert_headings(text)
    text = _convert_bullets(text)
    return md_to_telegram(text)


def format_for_qwen(text: str) -> str:
    """Format for Qwen3 Coder - uses **bold**, *italic*, `code`, ```code```"""
    if not text:
        return ""
    text = _convert_italic(text)
    text = _convert_bold(text)
    text = _convert_headings(text)
    text = _convert_bullets(text)
    return md_to_telegram(text)


def format_for_phi(text: str) -> str:
    """Format for Phi-4 - uses **bold**, *italic*, `code`, ```code```"""
    if not text:
        return ""
    text = _convert_italic(text)
    text = _convert_bold(text)
    text = _convert_headings(text)
    text = _convert_bullets(text)
    return md_to_telegram(text)


def format_for_minimax(text: str) -> str:
    """Format for MiniMax - uses **bold**, *italic*, `code`, ```code```"""
    if not text:
        return ""
    text = _convert_italic(text)
    text = _convert_bold(text)
    text = _convert_headings(text)
    text = _convert_bullets(text)
    return md_to_telegram(text)


def format_for_gpt_oss(text: str) -> str:
    """Format for GPT-OSS - uses **bold**, *italic*, `code`, ```code```"""
    if not text:
        return ""
    text = _convert_italic(text)
    text = _convert_bold(text)
    text = _convert_headings(text)
    text = _convert_bullets(text)
    return md_to_telegram(text)


# Map model keys to formatter functions
MODEL_FORMATTERS = {
    "google": format_for_gemma,
    "deepseek": format_for_deepseek,
    "kimi": format_for_kimi,
    "glm": format_for_glm,
    "nemotron": format_for_nemotron,
    "mistral": format_for_mistral,
    "mistral-medium": format_for_mistral,
    "llama": format_for_llama,
    "qwen": format_for_qwen,
    "phi": format_for_phi,
    "minimax": format_for_minimax,
    "gpt-oss": format_for_gpt_oss,
}


def format_for_model(text: str, model_key: str) -> str:
    """Format text for a specific model's markdown output style."""
    formatter = MODEL_FORMATTERS.get(model_key, md_to_telegram)
    return formatter(text)


def get_user_mention(user) -> str:
    name = user.first_name or user.username or "User"
    escaped_name = md_escape(name)
    return f"[{escaped_name}](tg://user?id={user.id})"


def split_text(text: str, max_len: int = 4000) -> list[str]:
    if len(text) <= max_len:
        return [text]

    parts = []
    current = ""

    paragraphs = text.split("\n\n")
    for para in paragraphs:
        if len(current) + len(para) + 2 <= max_len:
            current = f"{current}\n\n{para}" if current else para
        else:
            if current:
                parts.append(current)
            if len(para) <= max_len:
                current = para
            else:
                lines = para.split("\n")
                current = ""
                for line in lines:
                    if len(current) + len(line) + 1 <= max_len:
                        current = f"{current}\n{line}" if current else line
                    else:
                        if current:
                            parts.append(current)
                        if len(line) <= max_len:
                            current = line
                        else:
                            while len(line) > max_len:
                                parts.append(line[:max_len])
                                line = line[max_len:]
                            current = line

    if current:
        parts.append(current)

    return parts


def extract_files(text: str) -> tuple[list[tuple[str, str]], str]:
    """Extract [FILE: name]...[/FILE] blocks from AI response.
    Returns (list of (filename, content), remaining text)."""
    files = []
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")

    pattern = r'\[\s*FILE\s*:\s*([^\]\n]+?)\s*\]\s*\n?(.*?)\s*\[\s*/\s*FILE\s*\]'
    matches = list(re.finditer(pattern, normalized, re.DOTALL | re.IGNORECASE))

    for match in matches:
        filename = match.group(1).strip().strip("`\"' ")
        content = match.group(2)
        content = strip_code_block(content)
        if filename and content:
            files.append((filename, content))

    remaining = re.sub(pattern, '', normalized, flags=re.DOTALL | re.IGNORECASE).strip()

    if not files:
        loose_pattern = r'\[\s*FILE\s*:\s*([^\]\n]+?)\s*\]\s*\n?(.*?)(?=\n\[\s*FILE\s*:|\Z)'
        loose_matches = list(re.finditer(loose_pattern, normalized, re.DOTALL | re.IGNORECASE))
        if loose_matches:
            for match in loose_matches:
                filename = match.group(1).strip().strip("`\"' ")
                content = strip_code_block(match.group(2))
                if filename and content:
                    files.append((filename, content))
            remaining = re.sub(loose_pattern, '', normalized, flags=re.DOTALL | re.IGNORECASE).strip()

    return files, remaining


def strip_code_block(text: str) -> str:
    """Remove ```lang ... ``` wrapping from code (tolerant of missing closing fence)."""
    text = text.strip()
    if text.startswith("```"):
        newline_idx = text.find("\n")
        if newline_idx != -1:
            text = text[newline_idx + 1:]
        else:
            text = text[3:]
        if text.rstrip().endswith("```"):
            text = text.rstrip()[:-3]
    return text.strip()


def create_pdf(content: str, title: str = "Document") -> bytes:
    """Create a PDF from text content."""
    from fpdf import FPDF
    
    # Clean content - remove system-reminder tags and other artifacts
    content = re.sub(r'<system-reminder>.*?</system-reminder>', '', content, flags=re.DOTALL)
    content = re.sub(r'</?system-reminder[^>]*>', '', content)
    content = re.sub(r'</?thinking[^>]*>', '', content)
    content = content.strip()
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=20)
    
    # Set margins
    pdf.set_left_margin(15)
    pdf.set_right_margin(15)
    
    # Try to set a Unicode font
    font_set = False
    try:
        pdf.add_font("DejaVu", "", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", uni=True)
        pdf.set_font("DejaVu", size=11)
        font_set = True
    except Exception:
        pass
    
    if not font_set:
        try:
            pdf.add_font("Arial", "", "C:\\Windows\\Fonts\\arial.ttf", uni=True)
            pdf.set_font("Arial", size=11)
            font_set = True
        except Exception:
            pass
    
    if not font_set:
        pdf.set_font("Helvetica", size=11)
    
    # Title
    pdf.set_font_size(16)
    # Clean title of non-ASCII chars for built-in fonts
    if not font_set:
        title = title.encode("ascii", "replace").decode()
    pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)
    
    # Content
    pdf.set_font_size(11)
    
    for line in content.split("\n"):
        if not line.strip():
            pdf.ln(3)
            continue
        try:
            if not font_set:
                line = line.encode("ascii", "replace").decode()
            pdf.multi_cell(0, 6, line)
        except Exception:
            try:
                safe_line = line.encode("ascii", "replace").decode()
                pdf.multi_cell(0, 6, safe_line)
            except:
                continue
    
    return bytes(pdf.output())


async def send_files(message: Message, files: list[tuple[str, str]]):
    """Create and send files to the user."""
    for filename, content in files:
        safe_name = re.sub(r'[^\w\-.]', '_', filename)
        ext = os.path.splitext(safe_name)[1].lower()
        tmp_path = os.path.join(tempfile.gettempdir(), f"bot_{safe_name}")

        try:
            if ext == ".pdf":
                title = os.path.splitext(safe_name)[0]
                pdf_bytes = create_pdf(content, title)
                with open(tmp_path, 'wb') as f:
                    f.write(pdf_bytes)
            else:
                with open(tmp_path, 'w', encoding='utf-8') as f:
                    f.write(content)

            await message.reply_document(
                types.FSInputFile(tmp_path, filename=safe_name),
                caption=f"Created `{md_escape(safe_name)}`",
                parse_mode=ParseMode.MARKDOWN_V2,
            )
        except Exception as e:
            await message.reply(f"Failed to create `{md_escape(safe_name)}`: {md_escape(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


async def send_ai_response(message: Message, user_text: str, model_key: str, user_id: int, image_base64: str = None):
    cancel_kb = InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="Cancel", callback_data=CANCEL_AI)]]
    )
    
    if image_base64:
        thinking_msg = await message.reply(f"Analyzing image{_DOT}{_DOT}{_DOT}", reply_markup=cancel_kb, parse_mode=ParseMode.MARKDOWN_V2)
    else:
        thinking_msg = await message.reply(f"Thinking{_DOT}{_DOT}{_DOT}", reply_markup=cancel_kb, parse_mode=ParseMode.MARKDOWN_V2)
    
    req_id = thinking_msg.message_id
    active_requests[req_id] = {"user_id": user_id, "cancelled": False, "chat_id": message.chat.id}

    history = get_history(user_id)
    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    user_instruction = get_user_instruction(user_id)
    if user_instruction:
        messages.append({"role": "system", "content": f"User-specific instruction:\n{user_instruction}"})
    
    # Inject active agent content
    agent_content = get_active_agent_content(user_id)
    if agent_content:
        messages.append({"role": "system", "content": f"Active agent instructions:\n{agent_content}"})
    
    if is_question_mode(user_id):
        messages.append({"role": "system", "content": QUESTION_MODE_INSTRUCTION})
    messages.extend(history)
    messages.append({"role": "user", "content": user_text})

    clean_user_text = clean_ai_response(user_text)
    add_to_history(user_id, "user", clean_user_text)

    response = await call_ai(messages, model_key, image_base64)

    # If cancelled, don't send response
    is_cancelled = active_requests.get(req_id, {}).get("cancelled", False)
    active_requests.pop(req_id, None)

    if is_cancelled:
        try:
            await thinking_msg.delete()
        except Exception:
            pass
        return

    # Clean response BEFORE saving to history
    response = clean_ai_response(response)
    add_to_history(user_id, "assistant", response)

    # Question Mode: if the AI asked a clarifying question, render it with
    # inline answer buttons instead of sending a normal reply.
    if is_question_mode(user_id):
        question, options, preamble = parse_ask_block(response)
        if question:
            kb = None
            if options:
                kb = InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text=opt, callback_data=f"{QM_ANSWER_PREFIX}{i}")]
                    for i, opt in enumerate(options)
                ])
            pending_questions[user_id] = {"chat_id": message.chat.id, "options": options}
            try:
                await thinking_msg.delete()
            except Exception:
                pass
            qtext = f"{preamble}\n\n{question}".strip() if preamble else question
            await message.reply(qtext, reply_markup=kb)
            return

    try:
        await thinking_msg.delete()
    except Exception:
        pass

    files, remaining = extract_files(response)

    if files:
        await send_files(message, files)

    if not remaining and files:
        return

    if not remaining:
        remaining = response

    response = format_for_model(remaining, model_key)

    mention = ""
    if message.from_user:
        mention = get_user_mention(message.from_user)

    parts = split_text(response, max_len=4000)

    for i, part in enumerate(parts):
        if i == 0 and mention:
            text = f"{mention}\n\n{part}"
        else:
            text = part

        try:
            if i == 0:
                await message.reply(text, parse_mode=ParseMode.MARKDOWN_V2)
            else:
                await message.answer(text, parse_mode=ParseMode.MARKDOWN_V2)
        except Exception as e:
            logger.warning(f"MarkdownV2 parse failed, falling back to plain text: {e}")
            plain = re.sub(r'\\([_*\[\]()~`>#+\-=|{}.!\\])', r'\1', part)
            if i == 0:
                await message.reply(plain)
            else:
                await message.answer(plain)


# ─── Cancel Callback ─────────────────────────────────────────────────────────


@router.callback_query(F.data == CANCEL_AI)
async def cb_cancel_ai(callback: CallbackQuery):
    req_id = callback.message.message_id
    req = active_requests.get(req_id)
    if not req:
        await callback.answer("No active request found.", show_alert=True)
        return
    if req["user_id"] != callback.from_user.id and not is_owner(callback.from_user.id):
        await callback.answer("Only the user who asked or the owner can cancel this.", show_alert=True)
        return
    req["cancelled"] = True
    await callback.answer("Request cancelled.")
    try:
        await callback.message.delete()
    except Exception:
        pass


# ─── Command Handlers ────────────────────────────────────────────────────────


@router.message(CommandStart())
async def cmd_start(message: Message):
    # Check if started with /start agent parameter (from deep link)
    if message.text and len(message.text.split()) > 1 and message.text.split()[1] == "agent":
        await cmd_agent(message)
        return
    
    text = (
        f"Welcome{_EXC} I'm your *AI Group Manager Bot*{_DOT}\n\n"
        f"Add me to a group and use /authorize to activate me there{_DOT}\n"
        f"Use /settings to change the AI model{_DOT}\n"
        f"Use /q to ask me anything{_DOT}\n\n"
        f"> contact:@itznik\\_x"
    )
    await message.answer(text, parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("authorize","auth"))
async def cmd_authorize(message: Message):
    user_id = message.from_user.id
    
    if not is_owner(user_id):
        await message.answer(f"Only the bot owner can authorize{_DOT}", parse_mode=ParseMode.HTML)
        return
    
    command_parts = message.text.split() if message.text else []
    target_arg = command_parts[1] if len(command_parts) > 1 else None
    
    # If replying to a user's message, authorize that user globally
    if message.reply_to_message and message.reply_to_message.from_user:
        target = message.reply_to_message.from_user
        if target.id == bot.id:
            await message.answer(f"Cannot authorize the bot itself{_DOT}", parse_mode=ParseMode.HTML)
            return
authorize_user(target.id, user_id)
        name = target.first_name or target.username or str(target.id)
        await message.answer(
            f"User <b>{escape_html(name)}</b> (ID: <code>{target.id}</code>) has been globally authorized{_EXC}\n"
            f"They can now use the bot in DM and any group where I'm present{_EXC}",
            parse_mode=ParseMode.HTML,
        )
        return
    
    # Parse argument for user or group
    if target_arg:
        # Try to parse as username (with or without @)
        if target_arg.startswith("@") or (target_arg.replace("_", "").isalnum() and not target_arg.isdigit()):
            username = target_arg.lstrip("@")
            try:
                chat = await bot.get_chat(username)
                if chat.type == ChatType.PRIVATE:
                    if chat.id == bot.id:
                        await message.answer(f"Cannot authorize the bot itself{_DOT}", parse_mode=ParseMode.HTML)
                        return
                    authorize_user(chat.id, user_id)
                    name = chat.first_name or chat.username or str(chat.id)
                    await message.answer(
                        f"User <b>{escape_html(name)}</b> (ID: <code>{chat.id}</code>) has been globally authorized{_EXC}\n"
                        f"They can now use the bot in DM and any group where I'm present{_EXC}",
                        parse_mode=ParseMode.HTML,
                    )
                else:
                    await message.answer(f"Usernames only work for users, not groups/channels{_DOT}", parse_mode=ParseMode.HTML)
                return
            except Exception as e:
                await message.answer(f"Could not resolve username <code>{md_escape(username)}</code>{_DOT} {md_escape(str(e))}", parse_mode=ParseMode.HTML)
                return
        
# Try numeric argument
        if target_arg.lstrip("-").isdigit():
            target_id = int(target_arg)
            
            # If positive ID, try to resolve as user first (could be group)
            if target_id > 0:
                try:
                    chat = await bot.get_chat(target_id)
                    if chat.type == ChatType.PRIVATE:
                        if chat.id == bot.id:
                            await message.answer(f"Cannot authorize the bot itself{_DOT}", parse_mode=ParseMode.HTML)
                            return
                        authorize_user(chat.id, user_id)
                        name = chat.first_name or chat.username or str(chat.id)
                        await message.answer(
                            f"User <b>{escape_html(name)}</b> (ID: <code>{chat.id}</code>) has been globally authorized{_EXC}\n"
                            f"They can now use the bot in DM and any group where I'm present{_EXC}",
                            parse_mode=ParseMode.HTML,
                        )
                    else:
                        # It's a group/channel
                        if message.chat.type == ChatType.PRIVATE:
                            chat_id_str = str(target_id)
                            if chat_id_str in db["authorized_groups"]:
                                await message.answer(f"Group is already authorized{_DOT}", parse_mode=ParseMode.HTML)
                                return
                            
                            db["authorized_groups"][chat_id_str] = {
                                "title": f"Group {target_id}",
                                "authorized_at": datetime.now().isoformat(),
                                "authorized_by": user_id,
                            }
                            db["group_models"][chat_id_str] = DEFAULT_MODEL
                            save_data(db)
                            
                            await message.answer(
                                f"Group <code>{target_id}</code> has been authorized{_EXC}\n"
                                f"Current AI model: <b>{escape_html(MODELS[DEFAULT_MODEL]['name'])}</b>",
                                parse_mode=ParseMode.HTML,
                            )
                        else:
                            await message.answer(f"Use <code>/authorize</code> in DM to authorize groups by ID{_DOT}", parse_mode=ParseMode.HTML)
                        return
                except Exception:
                    # If get_chat fails, assume it's a user ID
                    authorize_user(target_id, user_id)
                    await message.answer(
                        f"User <code>{target_id}</code> has been globally authorized{_EXC}\n"
                        f"They can now use the bot in DM and any group where I'm present{_EXC}",
                        parse_mode=ParseMode.HTML,
                    )
                return
            
            # Negative ID = group
            else:
                if message.chat.type == ChatType.PRIVATE:
                    group_id = abs(target_id)
                    chat_id_str = str(group_id)
                    if chat_id_str in db["authorized_groups"]:
                        await message.answer(f"Group is already authorized{_DOT}", parse_mode=ParseMode.HTML)
                        return
                    
                    db["authorized_groups"][chat_id_str] = {
                        "title": f"Group {group_id}",
                        "authorized_at": datetime.now().isoformat(),
                        "authorized_by": user_id,
                    }
                    db["group_models"][chat_id_str] = DEFAULT_MODEL
                    save_data(db)
                    
                    await message.answer(
                        f"Group <code>{group_id}</code> has been authorized{_EXC}\n"
                        f"Current AI model: <b>{escape_html(MODELS[DEFAULT_MODEL]['name'])}</b>",
                        parse_mode=ParseMode.HTML,
                    )
                else:
                    await message.answer(f"Use <code>/authorize</code> in DM to authorize groups by ID{_DOT}", parse_mode=ParseMode.HTML)
                return
        
        await message.answer(f"Invalid argument{_DOT} Use <code>/authorize @username</code> or <code>/authorize <user_id></code> or reply to user{_DOT}", parse_mode=ParseMode.HTML)
        return
    
    # In a group - authorize the group itself
    if message.chat.type != ChatType.PRIVATE:
        chat_id = str(message.chat.id)
        chat_title = escape_html(message.chat.title or chat_id)
        
        if chat_id in db["authorized_groups"]:
            await message.answer(f"Group <b>{chat_title}</b> is already authorized{_DOT}", parse_mode=ParseMode.HTML)
            return
        
        db["authorized_groups"][chat_id] = {
            "title": message.chat.title or chat_id,
            "authorized_at": datetime.now().isoformat(),
            "authorized_by": user_id,
        }
        db["group_models"][chat_id] = DEFAULT_MODEL
        save_data(db)
        
        model_name = escape_html(MODELS[DEFAULT_MODEL]["name"])
        await message.answer(
            f"Group <b>{chat_title}</b> has been authorized{_EXC}\n"
            f"Current AI model: <b>{model_name}</b>\n"
            f"Use /settings to change the model{_DOT}",
            parse_mode=ParseMode.HTML,
        )
        return
    
    await message.answer(
        f"Usage:\n"
        f"- `/authorize` in a group - Authorize the group\n"
        f"- `/authorize <group_id>` in DM - Authorize a group by ID\n"
        f"- Reply to a user with `/authorize` - Globally authorize that user",
        parse_mode=ParseMode.MARKDOWN_V2,
    )


@router.message(Command("deauthorize", "unauthorize"))
async def cmd_deauthorize(message: Message):
    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can deauthorize{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    command_parts = message.text.split() if message.text else []
    target_arg = command_parts[1] if len(command_parts) > 1 else None
    
    # Reply to a user to deauthorize them globally
# If replying to a user to deauthorize them globally
    if message.reply_to_message and message.reply_to_message.from_user:
        target = message.reply_to_message.from_user
        if target.id == bot.id:
            await message.answer(f"Cannot deauthorize the bot{_DOT}", parse_mode=ParseMode.HTML)
            return
        deauthorize_user(target.id)
        name = target.first_name or target.username or str(target.id)
        await message.answer(
            f"User <b>{md_escape(name)}</b> has been globally deauthorized{_DOT}",
            parse_mode=ParseMode.HTML,
        )
        return
    
    # If target argument provided
    if target_arg:
        # Try username (with or without @)
        if target_arg.startswith("@") or (target_arg.replace(".", "").isalnum() and not target_arg.isdigit()):
            # This is likely a username, try to resolve it
            # We need to check if user is in authorized_users
            username = target_arg.lstrip("@").lower()
            found = False
            for uid, data in db.get("authorized_users", {}).items():
                # We don't store username, so we can't easily lookup by username
                # This would require a user lookup, which we can't do without API
                pass
            # Fall through to user ID check
        
        # Try user ID
        try:
            user_id = int(target_arg)
            if deauthorize_user(user_id):
                await message.answer(
                    f"User <code>{md_escape(str(user_id))}</code> has been globally deauthorized{_DOT}",
                    parse_mode=ParseMode.HTML,
                )
            else:
                await message.answer(f"User <code>{md_escape(str(user_id))}</code> was not authorized{_DOT}", parse_mode=ParseMode.HTML)
            return
        except ValueError:
            pass
    
    # In DM with a group ID
    if message.chat.type == ChatType.PRIVATE and target_arg:
        try:
            group_id = int(target_arg)
            chat_id_str = str(group_id)
            
            if chat_id_str not in db["authorized_groups"]:
                await message.answer(f"Group <code>{group_id}</code> is not authorized{_DOT}", parse_mode=ParseMode.HTML)
                return
            
            del db["authorized_groups"][chat_id_str]
            db["group_models"].pop(chat_id_str, None)
            db["conversations"].pop(chat_id_str, None)
            save_data(db)
            
            await message.answer(f"Group <code>{group_id}</code> has been deauthorized{_DOT}", parse_mode=ParseMode.HTML)
            return
        except ValueError:
            pass
    
    # In a group - deauthorize the group itself
    if message.chat.type != ChatType.PRIVATE:
        chat_id = str(message.chat.id)
        chat_title = md_escape(message.chat.title or chat_id)
        
        if chat_id not in db["authorized_groups"]:
            await message.answer(f"Group <b>{chat_title}</b> is not authorized{_DOT}", parse_mode=ParseMode.HTML)
            return
        
        del db["authorized_groups"][chat_id]
        db["group_models"].pop(chat_id, None)
        db["conversations"].pop(chat_id, None)
        save_data(db)
        
        await message.answer(f"Group <b>{chat_title}</b> has been deauthorized{_DOT}", parse_mode=ParseMode.HTML)
        return
    
    await message.answer(
        f"Usage:\n"
        f"- `/deauthorize` in a group - Deauthorize the group\n"
        f"- `/deauthorize <group_id>` in DM - Deauthorize a group by ID\n"
        f"- Reply to a user with `/deauthorize` - Globally deauthorize that user",
        parse_mode=ParseMode.MARKDOWN_V2,
    )


@router.message(Command("settings", "s"))
async def cmd_settings(message: Message):
    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can change settings{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    current_model = get_group_model(message.chat.id)
    current_name = md_escape(MODELS[current_model]["name"])

    await message.answer(
        f"Current model: *{current_name}* \\(global\\)\n\nSelect a new AI model:",
        reply_markup=model_keyboard(),
        parse_mode=ParseMode.MARKDOWN_V2,
    )


@router.callback_query(F.data.startswith("setmodel:"))
async def cb_set_model(callback: CallbackQuery):
    if not is_owner(callback.from_user.id):
        await callback.answer("Only the owner can change models.", show_alert=True)
        return

    model_key = callback.data.split(":")[1]
    if model_key not in MODELS:
        await callback.answer("Invalid model.", show_alert=True)
        return

    db["global_model"] = model_key
    db["group_models"] = {}
    save_data(db)

    model_name = md_escape(MODELS[model_key]["name"])
    await callback.message.edit_text(
        f"AI model changed to *{model_name}* \\(applied globally\\)",
        parse_mode=ParseMode.MARKDOWN_V2,
    )
    await callback.answer(f"Model set to {MODELS[model_key]['name']}")


@router.callback_query(F.data.startswith("modelpage:"))
async def cb_model_page(callback: CallbackQuery):
    if not is_owner(callback.from_user.id):
        await callback.answer("Only the owner can change models.", show_alert=True)
        return

    payload = callback.data.split(":", 1)[1]
    if payload == "noop":
        await callback.answer()
        return

    try:
        page = int(payload)
    except ValueError:
        await callback.answer("Invalid page.", show_alert=True)
        return

    try:
        await callback.message.edit_reply_markup(reply_markup=model_keyboard(page))
    except Exception:
        pass
    await callback.answer()


@router.message(Command("setname"))
async def cmd_setname(message: Message):
    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can change the bot name{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    command_text = (message.text or "").strip()
    if not is_setname_enabled() and not command_text.lower().endswith("enable"):
        await message.answer(
            f"Bot alias settings are disabled{_DOT}\n"
            f"Use `{md_escape('/setname enable')}` to enable them again{_DOT}",
            parse_mode=ParseMode.MARKDOWN_V2,
        )
        return

    if not is_setname_enabled() and command_text.lower().endswith("enable"):
        db["setname_enabled"] = True
        save_data(db)
        logger.info("Bot alias settings enabled by owner %s", message.from_user.id)

    alias = get_saved_bot_alias()
    status = "enabled" if is_setname_enabled() else "disabled"
    alias_text = md_escape(alias or "Not set")
    await message.answer(
        f"Bot alias settings: *{status}*{_DOT}\n"
        f"Bot alias name: *{alias_text}*{_DOT}\n\n"
        f"Use an alias like `{md_escape('babu')}` anywhere in a sentence{_DOT}\n"
        f"Choose an option below:",
        reply_markup=alias_keyboard(),
        parse_mode=ParseMode.MARKDOWN_V2,
    )


@router.callback_query(F.data == SETNAME_SET)
async def cb_setname_set(callback: CallbackQuery, state: FSMContext):
    if not is_owner(callback.from_user.id):
        await callback.answer("Only the owner can set the bot name.", show_alert=True)
        return

    if not is_setname_enabled():
        await callback.answer("Bot alias settings are disabled.", show_alert=True)
        return

    await state.set_state(SetNameState.waiting_alias)
    await state.set_data({"action": "set"})
    await callback.answer("Send the new bot alias name now.")
    if callback.message:
        await callback.message.edit_text(
            f"Send the new bot alias name now{_DOT}\n"
            f"Example: `{md_escape('yasir')}`",
            parse_mode=ParseMode.MARKDOWN_V2,
        )


@router.callback_query(F.data == SETNAME_EDIT)
async def cb_setname_edit(callback: CallbackQuery, state: FSMContext):
    if not is_owner(callback.from_user.id):
        await callback.answer("Only the owner can edit the bot name.", show_alert=True)
        return

    if not is_setname_enabled():
        await callback.answer("Bot alias settings are disabled.", show_alert=True)
        return

    alias = get_saved_bot_alias()
    if not alias:
        await callback.answer("No alias is set. Use Set name first.", show_alert=True)
        return

    await state.set_state(SetNameState.waiting_alias)
    await state.set_data({"action": "edit"})
    await callback.answer("Send the new bot alias name now.")
    if callback.message:
        await callback.message.edit_text(
            f"Current alias: *{md_escape(alias)}*{_DOT}\n"
            f"Send the new bot alias name now{_DOT}",
            parse_mode=ParseMode.MARKDOWN_V2,
        )


@router.callback_query(F.data == SETNAME_REMOVE)
async def cb_setname_remove(callback: CallbackQuery, state: FSMContext):
    if not is_owner(callback.from_user.id):
        await callback.answer("Only the owner can remove the bot name.", show_alert=True)
        return

    if not is_setname_enabled():
        await callback.answer("Bot alias settings are disabled.", show_alert=True)
        return

    await state.clear()
    db["bot_alias"] = None
    save_data(db)
    logger.info("Bot alias removed by owner %s", callback.from_user.id)
    await callback.answer("Bot alias removed.")
    if callback.message:
        await callback.message.edit_text(
            "Bot alias removed{_DOT}",
            parse_mode=ParseMode.MARKDOWN_V2,
        )


@router.callback_query(F.data == SETNAME_DISABLE)
async def cb_setname_disable(callback: CallbackQuery, state: FSMContext):
    if not is_owner(callback.from_user.id):
        await callback.answer("Only the owner can disable bot alias settings.", show_alert=True)
        return

    await state.clear()
    db["setname_enabled"] = False
    save_data(db)
    logger.info("Bot alias settings disabled by owner %s", callback.from_user.id)
    await callback.answer("Bot alias settings disabled.")
    if callback.message:
        await callback.message.edit_text(
            f"Bot alias settings disabled{_DOT}\n"
            f"Alias messages will be ignored until you use `{md_escape('/setname enable')}`{_DOT}",
            parse_mode=ParseMode.MARKDOWN_V2,
        )


@router.callback_query(F.data == SETNAME_ENABLE)
async def cb_setname_enable(callback: CallbackQuery, state: FSMContext):
    if not is_owner(callback.from_user.id):
        await callback.answer("Only the owner can enable bot alias settings.", show_alert=True)
        return

    db["setname_enabled"] = True
    save_data(db)
    logger.info("Bot alias settings enabled by owner %s", callback.from_user.id)
    await callback.answer("Bot alias settings enabled.")
    if callback.message:
        alias = get_saved_bot_alias()
        alias_text = md_escape(alias or "Not set")
        await callback.message.edit_text(
            f"Bot alias settings enabled{_DOT}\n"
            f"Current alias: *{alias_text}*{_DOT}\n\n"
            f"Choose an option below:",
            reply_markup=alias_keyboard(),
            parse_mode=ParseMode.MARKDOWN_V2,
        )


@router.message(Command("instructions", "i"))
async def cmd_instructions(message: Message, state: FSMContext):
    await state.clear()
    current = get_user_instruction(message.from_user.id)
    current_text = "Set" if current else "Not set"
    await message.answer(
        f"Your instruction: *{md_escape(current_text)}*{_DOT}\n\n"
        f"Choose an option below:",
        reply_markup=instruction_keyboard(),
        parse_mode=ParseMode.MARKDOWN_V2,
    )


@router.callback_query(F.data == INSTRUCTION_SET)
async def cb_instruction_set(callback: CallbackQuery, state: FSMContext):
    await state.set_state(InstructionState.waiting_instruction)
    await callback.answer("Send your instruction now.")
    if callback.message:
        await callback.message.edit_text(
            f"Send your personal instruction now{_DOT}\n"
            f"It will apply only to your messages{_DOT}",
            parse_mode=ParseMode.MARKDOWN_V2,
        )


@router.callback_query(F.data == INSTRUCTION_SHOW)
async def cb_instruction_show(callback: CallbackQuery):
    instruction = get_user_instruction(callback.from_user.id)
    if not instruction:
        await callback.answer("You have no instruction set.", show_alert=True)
        return

    text = md_to_telegram(f"*Your instruction:*\n\n{instruction}")
    if callback.message:
        try:
            await callback.message.answer(text, parse_mode=ParseMode.MARKDOWN_V2)
        except Exception:
            await callback.message.answer(instruction)
    await callback.answer("Instruction shown.")


@router.callback_query(F.data == INSTRUCTION_REMOVE)
async def cb_instruction_remove(callback: CallbackQuery):
    if callback.message and callback.message.chat.type != ChatType.PRIVATE and not is_authorized(callback.message.chat.id, callback.from_user.id):
        await callback.answer("This group is not authorized.", show_alert=True)
        return

    remove_user_instruction(callback.from_user.id)
    await callback.answer("Your instruction removed.")
    if callback.message:
        await callback.message.edit_text(
            f"Your instruction has been removed{_DOT}",
            parse_mode=ParseMode.MARKDOWN_V2,
        )


@router.message(InstructionState.waiting_instruction)
async def handle_instruction_input(message: Message, state: FSMContext):
    if not message.text:
        await state.clear()
        return

    if message.chat.type != ChatType.PRIVATE and not is_authorized(message.chat.id, message.from_user.id):
        await state.clear()
        await message.answer(f"This group is not authorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    if message.text.startswith("/"):
        await state.clear()
        if message.text.lower().startswith("/cancel"):
            await message.answer(f"Instruction setup cancelled{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    instruction = message.text.strip()
    if not instruction:
        await message.answer(f"Instruction cannot be empty{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    set_user_instruction(message.from_user.id, instruction)
    await state.clear()
    await message.answer(
        f"Your instruction has been saved{_DOT}\n"
        f"It will apply only to your messages{_DOT}",
        parse_mode=ParseMode.MARKDOWN_V2,
    )


@router.message(Command("setchannel"))
async def cmd_setchannel(message: Message):
    """Set the channel ID for image uploads."""
    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can use this command{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return
    
    if message.chat.type != ChatType.PRIVATE:
        await message.answer(f"Use this command in DM with the bot{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return
    
    if len(message.command) < 2:
        current = IMAGE_CHANNEL_ID or "Not set"
        await message.answer(
            f"*Image Channel Settings*\n\n"
            f"Current channel: `{current}`\n\n"
            f"Usage: `/setchannel \\-1001234567890`\n\n"
            f"To get channel ID:\n"
            f"{_DASH} Forward a message from the channel to @userinfobot\n"
            f"{_DASH} Or use /id in the channel\n\n"
            f"Make sure the bot is admin in the channel{_DOT}",
            parse_mode=ParseMode.MARKDOWN_V2,
        )
        return
    
    try:
        channel_id = int(message.command[1].strip())
    except ValueError:
        await message.answer(f"Invalid channel ID{_DOT} It should be a number like \\-1001234567890{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return
    
    # Test if bot can send to the channel
    try:
        test_msg = await bot.send_message(channel_id, "Testing channel connection...")
        await test_msg.delete()
        save_image_channel_id(channel_id)
        await message.answer(f"Image channel set to `{channel_id}`{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
    except Exception as e:
        await message.answer(f"Cannot send to channel `{channel_id}`{_DOT}\nMake sure the bot is admin there{_DOT}\nError: {md_escape(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("clearhistory","clr"))
async def cmd_clearhistory(message: Message):
    # Any user can clear their own history
    clear_history(message.from_user.id)
    await message.answer(f"Your conversation history cleared{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("clearallhistory","clrall"))
async def cmd_clearallhistory(message: Message):
    # Only owner can clear everyone's history
    if not is_owner(message.from_user.id):
        await message.answer(f"Only the bot owner can clear all history{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return
    
    db["conversations"] = {}
    save_data(db)
    await message.answer(f"All conversation history cleared{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("restart", "reboot","r"))
async def cmd_restart(message: Message):
    if not is_owner(message.from_user.id):
        await message.answer(
            f"Only the bot owner can restart the bot{_DOT}",
            parse_mode=ParseMode.MARKDOWN_V2,
        )
        return

    try:
        save_data(db)
    except Exception:
        pass

    for req_id in list(active_requests.keys()):
        active_requests[req_id]["cancelled"] = True

    notice = await message.answer(
        f"♻️ Restarting bot{_DOT}{_DOT}{_DOT} back in a moment{_DOT}",
        parse_mode=ParseMode.MARKDOWN_V2,
    )

    try:
        marker_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".restart_marker.json")
        with open(marker_path, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "chat_id": message.chat.id,
                    "notice_message_id": notice.message_id,
                    "user_id": message.from_user.id,
                    "started_at": datetime.utcnow().isoformat() + "Z",
                },
                f,
            )
    except Exception as e:
        logger.warning(f"Failed to write restart marker: {e}")

    try:
        await bot.session.close()
    except Exception:
        pass

    logger.info("Restart requested by owner — re-executing process.")

    async def _do_restart():
        await asyncio.sleep(0.5)
        os.execv(sys.executable, [sys.executable, os.path.abspath(__file__)] + sys.argv[1:])

    asyncio.create_task(_do_restart())


@router.message(Command("model"))
async def cmd_model(message: Message):
    if message.chat.type == ChatType.PRIVATE:
        await message.answer(f"This command only works in groups{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    if not is_authorized(message.chat.id, message.from_user.id):
        await message.answer(f"This group is not authorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    current = get_group_model(message.chat.id)
    name = md_escape(MODELS[current]["name"])
    await message.answer(f"Current AI model: *{name}*", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("q", "question"))
async def cmd_question(message: Message):
    if not message.text:
        return
    question_text = message.text.split(None, 1)
    if len(question_text) < 2 or not question_text[1].strip():
        await message.answer("Usage: `/q your question here`", parse_mode=ParseMode.MARKDOWN_V2)
        return

    user_text = question_text[1].strip()

    image_base64 = None

    if message.reply_to_message:
        replied = message.reply_to_message

        if replied.document:
            file_content = await download_and_read_file(message, document=replied.document)
            if file_content:
                file_name = replied.document.file_name or "file"
                user_text = f"User asked: {user_text}\n\n--- File: {file_name} ---\n{file_content}"

        elif replied.photo:
            image_base64 = await download_photo_as_base64(replied)
            user_text = f"User asked about this image: {user_text}"

        elif replied.text:
            replied_text = replied.text
            user_text = f"User asked: {user_text}\n\n--- Replied message ---\n{replied_text}"

        elif replied.caption:
            replied_text = replied.caption
            user_text = f"User asked: {user_text}\n\n--- Replied message ---\n{replied_text}"

    if message.chat.type == ChatType.PRIVATE:
        if not can_use_bot(message.from_user.id, message.chat.id):
            await reject_unauthorized(message)
            return
        await send_ai_response(message, user_text, DEFAULT_MODEL, message.from_user.id, image_base64)
        return

    if is_owner(message.from_user.id):
        model_key = get_group_model(message.chat.id)
        await send_ai_response(message, user_text, model_key, message.from_user.id, image_base64)
        return

    if not is_authorized_group(message.chat.id):
        await message.answer(f"This group is not authorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    model_key = get_group_model(message.chat.id)
    await send_ai_response(message, user_text, model_key, message.from_user.id, image_base64)


@router.message(Command("translate", "tr"))
async def cmd_translate(message: Message):
    if not message.text:
        return
    parts = message.text.split(None, 2)
    if len(parts) < 3:
        await message.answer(
            "<b>Usage:</b> <code>/translate french Hello how are you?</code>\n"
            "Languages: french, spanish, german, arabic, chinese, japanese, korean, hindi, portuguese, russian, italian, turkish",
            parse_mode=ParseMode.HTML,
        )
        return

    target_lang = parts[1].strip()
    text = parts[2].strip()

    thinking_msg = await message.reply(f"Translating to {target_lang}...", parse_mode=ParseMode.HTML)

    result = await call_translate(text, target_lang)

    try:
        await thinking_msg.delete()
    except Exception:
        pass

    safe_result = result.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    safe_lang = target_lang.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    await message.reply(
        f"<b>Translation to {safe_lang}:</b>\n\n{safe_result}",
        parse_mode=ParseMode.HTML,
    )


@router.message(Command("safety", "safe"))
async def cmd_safety(message: Message):
    text_to_check = None

    if message.reply_to_message and message.reply_to_message.text:
        text_to_check = message.reply_to_message.text
    elif message.text:
        parts = message.text.split(None, 1)
        if len(parts) >= 2 and parts[1].strip():
            text_to_check = parts[1].strip()

    if not text_to_check:
        await message.answer(
            "<b>Usage:</b> reply to a message with /safety or <code>/safety check this text</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    thinking_msg = await message.reply("Checking safety...", parse_mode=ParseMode.HTML)

    result = await call_safety(text_to_check)

    try:
        await thinking_msg.delete()
    except Exception:
        pass

    safe_result = result.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    if "unsafe" in result.lower():
        icon = "&#x26A0;"
    else:
        icon = "&#x2705;"
    await message.reply(
        f"<b>Safety Check:</b> {icon}\n\n{safe_result}",
        parse_mode=ParseMode.HTML,
    )


@router.message(Command("search", "websearch","find"))
async def cmd_search(message: Message):
    if not message.text:
        return
    parts = message.text.split(None, 1)
    if len(parts) < 2 or not parts[1].strip():
        await message.answer(
            "<b>Usage:</b> <code>/search your query here</code>\n\n"
            "Searches the web using Tavily and answers using the current AI model." \
            if False else f"Usage: `/search your query here`\n\nSearches the web using Tavily\\.",
            parse_mode=ParseMode.MARKDOWN_V2,
        )
        return

    query = parts[1].strip()

    if message.chat.type != ChatType.PRIVATE:
        if not is_authorized(message.chat.id, message.from_user.id):
            await message.answer(f"This group is not authorized{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
            return

    cancel_kb = InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="Cancel", callback_data=CANCEL_AI)]]
    )
    thinking_msg = await message.reply(
        f"Searching the web{_DOT}{_DOT}{_DOT}",
        reply_markup=cancel_kb,
        parse_mode=ParseMode.MARKDOWN_V2,
    )

    req_id = thinking_msg.message_id
    active_requests[req_id] = {
        "user_id": message.from_user.id,
        "cancelled": False,
        "chat_id": message.chat.id,
    }

    search_data = await web_search(query, max_results=5)

    if active_requests.get(req_id, {}).get("cancelled"):
        active_requests.pop(req_id, None)
        return

    if search_data.get("error") and not search_data.get("answer") and not search_data.get("results"):
        active_requests.pop(req_id, None)
        try:
            await thinking_msg.delete()
        except Exception:
            pass
        await message.reply(f"Search failed: {md_escape(search_data['error'])}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    search_context = format_search_results_for_ai(query, search_data)

    original_answer = search_data.get("answer", "")
    if not original_answer:
        results = search_data.get("results") or []
        if results:
            original_answer = results[0].get("content", "")[:500]
    augmented_query = (
        f"User asked: {query}\n\n"
        f"You have access to the following web search results. "
        f"Use them to give an accurate, up-to-date answer. "
        f"Cite sources naturally when relevant.\n\n"
        f"{search_context}\n\n"
        f"Now answer the user's question."
    )

    if message.chat.type == ChatType.PRIVATE:
        model_key = DEFAULT_MODEL
    else:
        model_key = get_group_model(message.chat.id)

    history = get_history(message.from_user.id)
    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    user_instruction = get_user_instruction(message.from_user.id)
    if user_instruction:
        messages.append({"role": "system", "content": f"User-specific instruction:\n{user_instruction}"})
    if history:
        trimmed = history[-4:]
        messages.extend(trimmed)
    messages.append({"role": "user", "content": augmented_query})

    clean_user_text = clean_ai_response(f"[Web Search] {query}")
    add_to_history(message.from_user.id, "user", clean_user_text)

    response = await call_ai(messages, model_key, enable_search=False)

    is_cancelled = active_requests.get(req_id, {}).get("cancelled", False)
    active_requests.pop(req_id, None)
    if is_cancelled:
        try:
            await thinking_msg.delete()
        except Exception:
            pass
        return

    response = clean_ai_response(response)
    add_to_history(message.from_user.id, "assistant", response)

    try:
        await thinking_msg.delete()
    except Exception:
        pass

    ai_body = format_for_model(response, model_key)

    parts_out = split_text(ai_body, max_len=4000)
    mention = ""
    if message.from_user:
        mention = get_user_mention(message.from_user)
    for i, part in enumerate(parts_out):
        if i == 0 and mention:
            text = f"{mention}\n\n{part}"
        else:
            text = part
        try:
            if i == 0:
                await message.reply(text, parse_mode=ParseMode.MARKDOWN_V2)
            else:
                await message.answer(text, parse_mode=ParseMode.MARKDOWN_V2)
        except Exception:
            if i == 0:
                await message.reply(part)
            else:
                await message.answer(part)

    ai_model_name = MODELS.get(model_key, {}).get("name", "")
    sources_html = format_search_sources_html(query, search_data, ai_model_name=ai_model_name)
    if sources_html:
        for chunk_start in range(0, len(sources_html), 3800):
            chunk = sources_html[chunk_start:chunk_start + 3800]
            try:
                await message.answer(
                    chunk,
                    parse_mode=ParseMode.HTML,
                    disable_web_page_preview=True,
                )
            except Exception:
                try:
                    plain = re.sub(r"<[^>]+>", "", chunk)
                    plain = plain.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
                    await message.answer(plain, disable_web_page_preview=True)
                except Exception:
                    pass


@router.message(Command("post"))
async def cmd_post(message: Message):
    content = None
    title = "Shared Content"
    image_url = None
    image_position = "above"  # Default: image above text
    
    # Parse command arguments for --below flag
    command_text = message.text or ""
    if "--below" in command_text:
        image_position = "below"
        command_text = command_text.replace("--below", "").strip()
    
    # Extract content from command arguments
    if command_text:
        parts = command_text.split(None, 1)
        if len(parts) >= 2 and parts[1].strip():
            content = parts[1].strip()
            title = content[:50].split("\n")[0] or "Shared Text"
    
    # Handle reply to message
    if message.reply_to_message:
        replied = message.reply_to_message
        
        # Handle photo reply
        if replied.photo:
            try:
                # Download photo
                photo = replied.photo[-1]  # Get largest size
                file = await bot.get_file(photo.file_id)
                
                with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
                    tmp_path = tmp.name
                    await bot.download_file(file.file_path, tmp_path)
                
                try:
                    # Upload to Telegraph
                    image_url = await upload_photo_to_telegraph(tmp_path)
                finally:
                    # Clean up temp file
                    try:
                        os.unlink(tmp_path)
                    except:
                        pass
                
                # Use caption as content if available
                if not content:
                    content = replied.caption or "Shared Image"
                    title = "Shared Image"
                    
            except Exception as e:
                logger.error(f"Failed to process photo: {e}")
                await message.reply(f"Failed to process photo: {md_escape(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)
                return
        
        # Handle document reply
        elif replied.document:
            if not content:
                content = await download_and_read_file(message, document=replied.document)
                title = replied.document.file_name or "File Content"
        
        # Handle text reply
        elif replied.text:
            if not content:
                content = replied.text
                title = replied.text[:50].split("\n")[0] or "Shared Text"
        
        # Handle caption reply
        elif replied.caption:
            if not content:
                content = replied.caption
                title = "Shared Content"
    
    # Check if we have content to publish
    if not content and not image_url:
        await message.answer(
            f"Usage: reply to a message with /post or /post your text here\n"
            f"Options: --below (place image below text)\n\n"
            f"Examples:\n"
            f"/post Hello World\n"
            f"/post --below Hello World\n"
            f"Reply to photo with /post\n"
            f"Reply to photo with /post --below My caption",
            parse_mode=ParseMode.MARKDOWN_V2,
        )
        return
    
    # Default content if only image
    if not content:
        content = "Shared Image"
    
    try:
        thinking = await message.reply(f"Publishing to Telegraph{_DOT}{_DOT}{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        
        # Create Telegraph page with or without image
        url = await create_telegraph_page(
            title=title,
            content=content,
            image_url=image_url,
            image_position=image_position
        )
        
        try:
            await thinking.delete()
        except Exception:
            pass
        
        author = message.from_user.first_name if message.from_user else "User"
        escaped_author = md_escape(author)
        escaped_url = md_escape(url)
        
        # Build success message
        success_msg = f"Published to Telegraph by *{escaped_author}*{_DOT}\n\n{escaped_url}"
        if image_url:
            if image_position == "above":
                success_msg += f"\n\n_Image above text_"
            else:
                success_msg += f"\n\n_Image below text_"
        
        await message.reply(success_msg, parse_mode=ParseMode.MARKDOWN_V2)
        
    except Exception as e:
        await message.reply(f"Failed to publish: {md_escape(str(e))}", parse_mode=ParseMode.MARKDOWN_V2)


@router.message(Command("help"))
async def cmd_help(message: Message):
    text = (
        "<b>AI Group Manager Bot - Commands</b>\n\n"
        "/start - Welcome message\n"
        "/q <code>question</code> - Ask AI a question directly\n"
        "/search <code>query</code> - Search the web with AI\n"
        "/translate <code>lang text</code> - Translate to another language\n"
        "/safety <code>text</code> - Check content safety\n"
        "/post - Publish text, code, or images to Telegraph\n"
        "/clearhistory - Clear your conversation history\n"
        "/instructions - Set your personal AI instruction\n"
        "/agent - Manage your AI agents (add / toggle / remove)\n"
        "/questionmode - Toggle clarifying-question mode (per-user)\n"
        "/instruction - Show the getting-started guide\n"
        "/model - Show current AI model\n"
        "/help - Show this help message\n\n"
        "<b>Owner Only Commands:</b>\n"
        "/authorize - Authorize group or user (reply to auth user globally)\n"
        "/deauthorize or /unauthorize - Deauthorize group or user\n"
        "/settings - Change AI model\n"
        "/setchannel - Set image channel for /post\n"
        "/setname - Set bot alias name\n"
        "/setname enable - Re-enable bot alias settings if disabled\n"
        "/clearallhistory - Clear everyone's history\n"
        "/log - Send bot logs since startup\n\n"
        "<b>How to use:</b>\n"
        "- Reply to my message to chat with me\n"
        "- Tag me with @ to ask something\n"
        "- Use /setname to set an alias like <code>yasir</code>, then send <code>yasir what is the day today</code>\n"
        "- Send any file with alias/mention to read it (txt, pdf, docx, xlsx, zip, etc)\n"
        "- Reply to any message or file with /q\n"
        "- Use /search to get answers from the live web with cited sources\n"
        "- Use /post to publish text or code to Telegraph\n"
        "- Use /post with --below flag for image below text\n"
        "- Reply to photos with /post to publish them\n"
        "- Owner can chat directly without /q in DM\n"
        "- Globally authorized users can use bot anywhere (DM + groups)\n"
        "- Cancel button appears during AI thinking; only you can cancel your request\n\n"
        "<b>Authorization (Leech Bot):</b>\n"
        "- Owner can use bot without authorization\n"
        "- Reply to a user with /authorize to grant global access\n"
        "- Users in an authorized group can use bot in that group only\n\n"
        "<b>Available Models:</b>\n"
        "Nemotron, DeepSeek, Mistral Large, Llama, Qwen3 Coder, Phi-4\n"
        "Kimi K2.6, Mistral Medium, MiniMax M2.7, GPT-OSS 120B"
    )
    await message.answer(text, parse_mode=ParseMode.HTML)


# ─── Document Handler (File Reading) ─────────────────────────────────────────


@router.message(AgentState.waiting_file, F.document)
async def handle_agent_file(message: Message, state: FSMContext):
    doc = message.document
    file_name = doc.file_name or "agent.md"

    if not file_name.lower().endswith(".md"):
        await message.answer("Only *.md* files are accepted. Please send a .md file.")
        return

    if doc.file_size and doc.file_size > MAX_AGENT_FILE_KB * 1024:
        await message.answer(f"File too large. Maximum size is *{MAX_AGENT_FILE_KB}KB*.")
        return

    file = await bot.get_file(doc.file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".md") as tmp:
        tmp_path = tmp.name
        await bot.download_file(file.file_path, tmp_path)

    try:
        with open(tmp_path, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception as e:
        await message.answer(f"Error reading file: {e}")
        os.unlink(tmp_path)
        return
    finally:
        try:
            os.unlink(tmp_path)
        except:
            pass

    await message.delete()
    await state.update_data(content=content, filename=file_name)

    buttons = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Default Name", callback_data=AGENT_RENAME_DEFAULT)],
        [InlineKeyboardButton(text="Change Name", callback_data=AGENT_RENAME_CUSTOM)],
    ])
    await message.answer(
        f"Do you want to rename your agent or keep the default filename *{md_escape(file_name)}*?",
        reply_markup=buttons,
        parse_mode=ParseMode.MARKDOWN_V2,
    )


@router.message(AgentState.waiting_rename)
async def handle_agent_rename(message: Message, state: FSMContext):
    if not message.text:
        return

    if message.text.startswith("/"):
        await state.clear()
        return

    name = message.text.strip()[:40]
    if not name:
        await message.answer("Name cannot be empty.")
        return

    data = await state.get_data()
    content = data.get("content", "")
    filename = data.get("filename", "agent.md")

    user_agents = get_user_agents(message.from_user.id)
    user_agents.append({
        "name": name,
        "filename": filename,
        "content": content,
        "active": True,
        "added_at": datetime.now().isoformat(),
    })
    save_user_agents(message.from_user.id, user_agents)
    await state.clear()

    escaped_name = md_escape(name)
    success_msg = await message.answer(f"Agent <b>{html.escape(name)}</b> saved successfully!", parse_mode=ParseMode.HTML)
    await asyncio.sleep(2)
    try:
        await success_msg.delete()
    except:
        pass
    active_count = sum(1 for a in user_agents if a.get("active", True))
    text = f"You have <b>{active_count} / {len(user_agents)}</b> agents active (max 5). Tap to toggle."
    await message.answer(text, reply_markup=agent_menu_keyboard(user_agents), parse_mode=ParseMode.HTML)


@router.message(F.document)
async def handle_document(message: Message):
    if message.chat.type == ChatType.PRIVATE:
        if not can_use_bot(message.from_user.id, message.chat.id):
            await reject_unauthorized(message)
            return

        file_content = await download_and_read_file(message)
        if file_content is None:
            await message.answer(f"File format not supported{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
            return

        caption = message.caption or "Summarize and explain the content of this file."
        user_text = f"{caption}\n\n--- File Content ---\n{file_content}"

        await send_ai_response(message, user_text, DEFAULT_MODEL, message.from_user.id)
        return

    # Group chat: only respond if explicitly addressed (reply-to-bot or @mention),
    # even for the owner — otherwise sharing files in the group triggers the bot randomly.
    is_reply_to_bot = (
        message.reply_to_message
        and message.reply_to_message.from_user
        and message.reply_to_message.from_user.id == bot.id
    )
    is_mentioned = is_bot_mentioned(message)

    if not is_reply_to_bot and not is_mentioned:
        return

    if is_reply_to_bot and mentions_other_user(message):
        return

    if not is_owner(message.from_user.id) and not can_use_bot(message.from_user.id, message.chat.id):
        await reject_unauthorized(message)
        return

    file_content = await download_and_read_file(message)
    if file_content is None:
        await message.answer(f"File format not supported{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    caption = message.caption or "Summarize and explain the content of this file."
    user_text = f"{caption}\n\n--- File Content ---\n{file_content}"

    model_key = get_group_model(message.chat.id)
    await send_ai_response(message, user_text, model_key, message.from_user.id)


@router.message(F.photo & F.caption)
async def handle_photo_with_caption(message: Message):
    caption = message.caption or ""
    if not caption.startswith("/q") and not caption.startswith("/question"):
        return

    if not can_use_bot(message.from_user.id, message.chat.id):
        await reject_unauthorized(message)
        return

    question = caption.split(None, 1)
    user_text = question[1].strip() if len(question) > 1 else "Describe this image in detail."

    thinking_msg = await message.reply(f"Analyzing image{_DOT}{_DOT}{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)

    image_base64 = await download_photo_as_base64(message)
    if not image_base64:
        try:
            await thinking_msg.delete()
        except:
            pass
        await message.answer(f"Failed to download image{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    model_key = get_group_model(message.chat.id)
    if not is_authorized(message.chat.id) and not is_owner(message.from_user.id):
        model_key = DEFAULT_MODEL

    history = get_history(message.from_user.id)
    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    user_instruction = get_user_instruction(message.from_user.id)
    if user_instruction:
        messages.append({"role": "system", "content": f"User-specific instruction:\n{user_instruction}"})
    messages.extend(history)
    messages.append({"role": "user", "content": user_text})

    add_to_history(message.from_user.id, "user", f"[Image] {user_text}")

    response = await call_ai(messages, model_key, image_base64)
    
    # Clean response BEFORE saving to history
    response = clean_ai_response(response)
    add_to_history(message.from_user.id, "assistant", response)

    try:
        await thinking_msg.delete()
    except:
        pass

    response = format_for_model(response, model_key)
    parts = split_text(response, max_len=4000)

    for i, part in enumerate(parts):
        try:
            if i == 0:
                await message.reply(part, parse_mode=ParseMode.MARKDOWN_V2)
            else:
                await message.answer(part, parse_mode=ParseMode.MARKDOWN_V2)
        except:
            if i == 0:
                await message.reply(part)
            else:
                await message.answer(part)


@router.message(F.photo)
async def handle_photo(message: Message):
    return


# ─── Message Handler (AI Reply) ──────────────────────────────────────────────


def extract_tagged_users(message: Message) -> str:
    """Extract tagged users from message and return context string."""
    tagged_users = []
    
    if message.entities:
        for entity in message.entities:
            if entity.type == "text_mention" and entity.user:
                # User without username
                name = entity.user.first_name or "User"
                tagged_users.append(f"@{name} (ID: {entity.user.id})")
            elif entity.type == "mention":
                # User with username
                mention_text = message.text[entity.offset:entity.offset + entity.length]
                tagged_users.append(mention_text)
    
    if message.reply_to_message and message.reply_to_message.from_user:
        replied_user = message.reply_to_message.from_user
        name = replied_user.first_name or "User"
        if replied_user.username:
            tagged_users.append(f"@{replied_user.username}")
        else:
            tagged_users.append(f"@{name} (ID: {replied_user.id})")
        
        # Also include replied message text
        if message.reply_to_message.text:
            return f"[Replying to {name}: \"{message.reply_to_message.text}\"]\n"
    
    if tagged_users:
        return f"[Tagged users: {', '.join(tagged_users)}]\n"
    
    return ""


@router.message(SetNameState.waiting_alias)
async def handle_alias_input(message: Message, state: FSMContext):
    if not message.text:
        await state.clear()
        return

    if not is_owner(message.from_user.id):
        await state.clear()
        await message.answer(f"Only the bot owner can set the alias name{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    if message.text.startswith("/"):
        await state.clear()
        if message.text.lower().startswith("/cancel"):
            await message.answer(f"Alias setup cancelled{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    alias = normalize_bot_alias(message.text)
    if not alias:
        await message.answer(f"Invalid alias name{_DOT} Send a non-empty name up to 40 characters{_DOT}", parse_mode=ParseMode.MARKDOWN_V2)
        return

    action = "updated" if get_bot_alias() else "set"
    db["bot_alias"] = alias
    save_data(db)
    logger.info("Bot alias %s to '%s' by owner %s", action, alias, message.from_user.id)

    question_text = strip_bot_alias(message.text)
    if question_text:
        model_key = DEFAULT_MODEL if message.chat.type == ChatType.PRIVATE else get_group_model(message.chat.id)
        await state.clear()
        await send_ai_response(message, question_text, model_key, message.from_user.id)
        return

    await state.clear()
    await message.answer(
        f"Bot alias {action} to *{md_escape(alias)}*{_DOT}\n"
        f"Now you can ask: `{md_escape(alias + ' what is the day today')}`",
        parse_mode=ParseMode.MARKDOWN_V2,
    )


@router.message(F.text & ~F.text.startswith("/"))
async def handle_message(message: Message):
    # Owner and globally-authorized users can chat freely in private messages
    if message.chat.type == ChatType.PRIVATE and can_use_bot(message.from_user.id, message.chat.id):
        user_text = message.text or message.caption or ""
        alias_text = strip_bot_alias(user_text)
        if alias_text is not None:
            user_text = alias_text
        if not user_text:
            return
        
        # Check for documents in DM (owner)
        if message.document:
            file_content = await download_and_read_file(message)
            if file_content:
                file_name = message.document.file_name or "file"
                user_text = f"{user_text}\n\n--- File: {file_name} ---\n{file_content}"
        elif message.photo:
            image_base64 = await download_photo_as_base64(message)
            if image_base64:
                model_key = DEFAULT_MODEL
                await send_ai_response(message, user_text, model_key, message.from_user.id, image_base64)
                return
        
        model_key = DEFAULT_MODEL
        await send_ai_response(message, user_text, model_key, message.from_user.id)
        return

    if message.chat.type == ChatType.PRIVATE:
        # Private chat, but not owner or globally authorized
        await reject_unauthorized(message)
        return

    # For groups: only respond to bot mentions, alias names, or replies to bot
    is_reply_to_bot = (
        message.reply_to_message
        and message.reply_to_message.from_user
        and message.reply_to_message.from_user.id == bot.id
    )
    is_mentioned = is_bot_mentioned(message)
    is_alias = is_alias_message(message)

    if not is_reply_to_bot and not is_mentioned and not is_alias:
        return

    if is_reply_to_bot and mentions_other_user(message):
        return

    if not can_use_bot(message.from_user.id, message.chat.id):
        await reject_unauthorized(message)
        return

    # Extract text from alias/mention
    if is_alias:
        user_text = strip_bot_alias(message.text or "")
    elif is_mentioned:
        user_text = strip_bot_mention(message.text or "")
    else:
        user_text = message.text or ""

    if not user_text:
        user_text = "What do you think about this?"

    # Check for documents/photos when alias/mention is triggered
    if message.document:
        file_content = await download_and_read_file(message)
        if file_content:
            file_name = message.document.file_name or "file"
            user_text = f"{user_text}\n\n--- File: {file_name} ---\n{file_content}"
    elif message.photo:
        image_base64 = await download_photo_as_base64(message)
        if image_base64:
            model_key = get_group_model(message.chat.id)
            await send_ai_response(message, user_text, model_key, message.from_user.id, image_base64)
            return
    
    # Add context about tagged users (if any)
    tagged_context = extract_tagged_users(message)
    if tagged_context:
        user_text = tagged_context + user_text

    model_key = get_group_model(message.chat.id)
    await send_ai_response(message, user_text, model_key, message.from_user.id)


# ─── Agent Command & Handlers ───────────────────────────────────────────────


def agent_menu_keyboard(user_agents: list) -> InlineKeyboardMarkup:
    buttons = []
    if user_agents:
        for i, agent in enumerate(user_agents):
            name = agent.get("name", agent.get("filename", "Unnamed"))
            active = agent.get("active", True)
            prefix = "✅ " if active else "⬜ "
            buttons.append([InlineKeyboardButton(text=f"{prefix}{name}", callback_data=f"{AGENT_TOGGLE_PREFIX}{i}")])
    buttons.append([
        InlineKeyboardButton(text="Add Agent", callback_data=AGENT_ADD),
        InlineKeyboardButton(text="Remove Agent", callback_data=AGENT_REMOVE),
    ])
    buttons.append([
        InlineKeyboardButton(text="Close", callback_data=AGENT_CLOSE),
    ])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def agent_remove_keyboard(user_agents: list, selected: set, page: int = 0) -> InlineKeyboardMarkup:
    buttons = []
    start = page * AGENTS_PER_PAGE
    end = min(start + AGENTS_PER_PAGE, len(user_agents))
    for i in range(start, end):
        name = user_agents[i].get("name", user_agents[i].get("filename", "Unnamed"))
        prefix = "✅ " if i in selected else "⬜ "
        buttons.append([InlineKeyboardButton(text=f"{prefix}{name}", callback_data=f"{AGENT_SELECT_PREFIX}{i}")])
    
    nav_buttons = []
    if page > 0:
        nav_buttons.append(InlineKeyboardButton(text="⬅", callback_data=f"{AGENT_PAGE_PREFIX}{page - 1}"))
    if end < len(user_agents):
        nav_buttons.append(InlineKeyboardButton(text="➡", callback_data=f"{AGENT_PAGE_PREFIX}{page + 1}"))
    if nav_buttons:
        buttons.append(nav_buttons)
    
    buttons.append([
        InlineKeyboardButton(text="Done", callback_data=AGENT_DONE),
        InlineKeyboardButton(text="Cancel", callback_data=AGENT_CANCEL),
    ])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


@router.message(Command("agent", "agents"))
async def cmd_agent(message: Message):
    if message.chat.type != ChatType.PRIVATE:
        bot_info = await bot.get_me()
        deep_link = f"https://t.me/{bot_info.username}?start=agent"
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Open DM", url=deep_link)]
        ])
        await message.answer(
            f"Agent management is private and only works in a direct chat with the bot{_DOT}\n"
            f"Please message me in private to manage your agents{_DOT}",
            reply_markup=kb,
            parse_mode=ParseMode.HTML,
        )
        return

    user_agents = get_user_agents(message.from_user.id)
    if not user_agents:
        text = "You don't have any current agents."
    else:
        active_count = sum(1 for a in user_agents if a.get("active", True))
        text = f"You have <b>{active_count} / {len(user_agents)}</b> agents active (max 5). Tap to toggle."

    await message.answer(text, reply_markup=agent_menu_keyboard(user_agents), parse_mode=ParseMode.HTML)


@router.callback_query(F.data.startswith(AGENT_TOGGLE_PREFIX))
async def cb_agent_toggle(callback: CallbackQuery):
    if not await _ensure_agent_private(callback):
        return
    try:
        idx = int(callback.data.split(":")[-1])
    except ValueError:
        await callback.answer("Invalid selection.")
        return

    user_agents = get_user_agents(callback.from_user.id)
    if not user_agents or idx >= len(user_agents):
        await callback.answer("Agent not found.")
        return

    current_active = user_agents[idx].get("active", True)
    active_count = sum(1 for a in user_agents if a.get("active", True))

    if current_active:
        user_agents[idx]["active"] = False
    else:
        if active_count >= 5:
            await callback.answer("Maximum 5 agents can be active at a time.", show_alert=True)
            return
        user_agents[idx]["active"] = True

    save_user_agents(callback.from_user.id, user_agents)
    await callback.answer()
    if callback.message:
        active_count = sum(1 for a in user_agents if a.get("active", True))
        text = f"You have <b>{active_count} / {len(user_agents)}</b> agents active (max 5). Tap to toggle."
        await callback.message.edit_text(text, reply_markup=agent_menu_keyboard(user_agents), parse_mode=ParseMode.HTML)


@router.callback_query(F.data == AGENT_CLOSE)
async def cb_agent_close(callback: CallbackQuery):
    if not await _ensure_agent_private(callback):
        return
    await callback.answer("Closed.")
    if callback.message:
        try:
            await callback.message.delete()
        except:
            pass


@router.callback_query(F.data == AGENT_ADD)
async def cb_agent_add(callback: CallbackQuery, state: FSMContext):
    if not await _ensure_agent_private(callback):
        return
    user_agents = get_user_agents(callback.from_user.id)
    if len(user_agents) >= MAX_AGENTS:
        await callback.answer(f"Maximum {MAX_AGENTS} agents allowed.", show_alert=True)
        return
    
    await state.set_state(AgentState.waiting_file)
    await callback.answer("Send your .md file now.")
    if callback.message:
        await callback.message.edit_text(
            f"Send your <b>.md</b> agent file.\n"
            f"Maximum size: <b>{MAX_AGENT_FILE_KB}KB</b>.\n"
            f"You can have up to <b>{MAX_AGENTS}</b> agents.",
            parse_mode=ParseMode.HTML,
        )



@router.callback_query(F.data == AGENT_RENAME_DEFAULT)
async def cb_agent_rename_default(callback: CallbackQuery, state: FSMContext):
    if not await _ensure_agent_private(callback):
        return
    data = await state.get_data()
    content = data.get("content", "")
    filename = data.get("filename", "agent.md")
    name = os.path.splitext(filename)[0].replace("_", " ").replace("-", " ").title()
    
    user_agents = get_user_agents(callback.from_user.id)
    user_agents.append({
        "name": name,
        "filename": filename,
        "content": content,
        "active": True,
        "added_at": datetime.now().isoformat(),
    })
    save_user_agents(callback.from_user.id, user_agents)
    await state.clear()

    await callback.answer("Agent saved successfully!")
    if callback.message:
        await callback.message.edit_text(f"Agent <b>{html.escape(name)}</b> saved successfully!", parse_mode=ParseMode.HTML)
        await asyncio.sleep(2)
        active_count = sum(1 for a in user_agents if a.get("active", True))
        text = f"You have <b>{active_count} / {len(user_agents)}</b> agents active (max 5). Tap to toggle."
        await callback.message.edit_text(text, reply_markup=agent_menu_keyboard(user_agents), parse_mode=ParseMode.HTML)


@router.callback_query(F.data == AGENT_RENAME_CUSTOM)
async def cb_agent_rename_custom(callback: CallbackQuery, state: FSMContext):
    if not await _ensure_agent_private(callback):
        return
    await state.set_state(AgentState.waiting_rename)
    await callback.answer("Enter the new name now.")
    if callback.message:
        await callback.message.edit_text(
            "Send the new name for your agent.\n"
            "Max 40 characters.",
            parse_mode=ParseMode.HTML,
        )



@router.callback_query(F.data == AGENT_REMOVE)
async def cb_agent_remove(callback: CallbackQuery, state: FSMContext):
    if not await _ensure_agent_private(callback):
        return
    user_agents = get_user_agents(callback.from_user.id)
    if not user_agents:
        await callback.answer("No agents to remove.", show_alert=True)
        return
    
    await state.update_data(selected=[], page=0)
    await callback.answer()
    if callback.message:
        await callback.message.edit_text(
            "Select agents to remove. When ready, tap Done.",
            reply_markup=agent_remove_keyboard(user_agents, selected=set()),
            parse_mode=ParseMode.HTML,
        )


@router.callback_query(F.data.startswith(AGENT_SELECT_PREFIX))
async def cb_agent_select(callback: CallbackQuery, state: FSMContext):
    if not await _ensure_agent_private(callback):
        return
    try:
        idx = int(callback.data.split(":")[-1])
    except ValueError:
        await callback.answer("Invalid selection.")
        return
    
    user_agents = get_user_agents(callback.from_user.id)
    if not user_agents or idx >= len(user_agents):
        await callback.answer("Agent not found.")
        return
    
    state_data = await state.get_data()
    selected = set(state_data.get("selected", []))
    
    if idx in selected:
        selected.remove(idx)
    else:
        if len(selected) >= 5:
            await callback.answer("Maximum 5 agents can be selected at a time.", show_alert=True)
            return
        selected.add(idx)
    
    await state.update_data(selected=list(selected))
    
    page = state_data.get("page", 0)
    
    if callback.message:
        await callback.message.edit_reply_markup(
            reply_markup=agent_remove_keyboard(user_agents, selected, page)
        )
    await callback.answer()


@router.callback_query(F.data.startswith(AGENT_PAGE_PREFIX))
async def cb_agent_page(callback: CallbackQuery, state: FSMContext):
    if not await _ensure_agent_private(callback):
        return
    try:
        page = int(callback.data.split(":")[-1])
    except ValueError:
        await callback.answer("Invalid page.")
        return
    
    user_agents = get_user_agents(callback.from_user.id)
    state_data = await state.get_data()
    selected = set(state_data.get("selected", []))
    await state.update_data(page=page)
    
    if callback.message:
        await callback.message.edit_reply_markup(
            reply_markup=agent_remove_keyboard(user_agents, selected, page)
        )
    await callback.answer()


@router.callback_query(F.data == AGENT_DONE)
async def cb_agent_done(callback: CallbackQuery, state: FSMContext):
    if not await _ensure_agent_private(callback):
        return
    state_data = await state.get_data()
    selected = set(state_data.get("selected", []))
    
    if not selected:
        await callback.answer("Please select at least one agent.", show_alert=True)
        return
    
    user_agents = get_user_agents(callback.from_user.id)
    # Remove in reverse order so indices remain valid
    for idx in sorted(selected, reverse=True):
        if 0 <= idx < len(user_agents):
            user_agents.pop(idx)
    
    save_user_agents(callback.from_user.id, user_agents)
    await state.clear()
    
    await callback.answer(f"Removed {len(selected)} agent(s).")
    if callback.message:
        if user_agents:
            names = "\n".join(f"{i+1}. {a.get('name', 'Unnamed')}" for i, a in enumerate(user_agents))
            await callback.message.edit_text(
                f"Agents removed. Current agents:\n{names}",
                reply_markup=agent_menu_keyboard(user_agents),
            )
        else:
            await callback.message.edit_text(
                "You don't have any current agents.",
                reply_markup=agent_menu_keyboard(user_agents),
            )


@router.callback_query(F.data == AGENT_CANCEL)
async def cb_agent_cancel(callback: CallbackQuery, state: FSMContext):
    if not await _ensure_agent_private(callback):
        return
    await state.clear()
    
    user_agents = get_user_agents(callback.from_user.id)
    await callback.answer("Cancelled.")
    if callback.message:
        if user_agents:
            names = "\n".join(f"{i+1}. {a.get('name', 'Unnamed')}" for i, a in enumerate(user_agents))
            await callback.message.edit_text(
                f"Operation cancelled. Your agents:\n{names}",
                reply_markup=agent_menu_keyboard(user_agents),
            )
        else:
            await callback.message.edit_text(
                "Operation cancelled. You don't have any current agents.",
                reply_markup=agent_menu_keyboard(user_agents),
            )


# ─── Question Mode Command & Handlers ───────────────────────────────────────


def _qm_menu_markup(enabled: bool, user_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(
                text=("✅ Enable" if enabled else "Enable"),
                callback_data=f"{QM_ENABLE}{user_id}",
            ),
            InlineKeyboardButton(
                text=("✅ Disable" if not enabled else "Disable"),
                callback_data=f"{QM_DISABLE}{user_id}",
            ),
        ],
        [
            InlineKeyboardButton(text="Close", callback_data=f"{QM_CLOSE}{user_id}"),
        ]
    ])


def _qm_menu_text(enabled: bool) -> str:
    return (
        f"<b>Question Mode</b> (currently {'ON' if enabled else 'OFF'}) — per-user setting\n\n"
        f"When ON, I ask clarifying questions with inline buttons for complex or "
        f"building/development queries, and I avoid guessing."
    )


@router.message(Command("questionmode"))
async def cmd_questionmode(message: Message):
    user_id = message.from_user.id
    enabled = is_question_mode(user_id)
    await message.answer(
        _qm_menu_text(enabled),
        reply_markup=_qm_menu_markup(enabled, user_id),
        parse_mode=ParseMode.HTML,
    )


@router.callback_query(F.data.startswith(QM_ENABLE))
async def cb_qm_enable(callback: CallbackQuery):
    try:
        target_user_id = int(callback.data[len(QM_ENABLE):])
    except (ValueError, IndexError):
        await callback.answer("Invalid request.", show_alert=True)
        return
    
    # Allow owner to change anyone's setting, users can only change their own
    if not (_is_owner(callback.from_user.id) or callback.from_user.id == target_user_id):
        await callback.answer("You can only change your own setting.", show_alert=True)
        return
    
    set_question_mode(target_user_id, True)
    await callback.message.edit_text(
        _qm_menu_text(True),
        reply_markup=_qm_menu_markup(True, target_user_id),
        parse_mode=ParseMode.HTML,
    )
    await callback.answer("Question Mode enabled.")


@router.callback_query(F.data.startswith(QM_DISABLE))
async def cb_qm_disable(callback: CallbackQuery):
    try:
        target_user_id = int(callback.data[len(QM_DISABLE):])
    except (ValueError, IndexError):
        await callback.answer("Invalid request.", show_alert=True)
        return
    
    if not (_is_owner(callback.from_user.id) or callback.from_user.id == target_user_id):
        await callback.answer("You can only change your own setting.", show_alert=True)
        return
    
    set_question_mode(target_user_id, False)
    await callback.message.edit_text(
        _qm_menu_text(False),
        reply_markup=_qm_menu_markup(False, target_user_id),
        parse_mode=ParseMode.HTML,
    )
    await callback.answer("Question Mode disabled.")


@router.callback_query(F.data.startswith(QM_CLOSE))
async def cb_qm_close(callback: CallbackQuery):
    try:
        target_user_id = int(callback.data[len(QM_CLOSE):])
    except (ValueError, IndexError):
        await callback.answer("Invalid request.", show_alert=True)
        return
    
    if not (_is_owner(callback.from_user.id) or callback.from_user.id == target_user_id):
        await callback.answer("You can only close your own menu.", show_alert=True)
        return
    
    try:
        await callback.message.delete()
    except Exception:
        pass
    await callback.answer("Closed.")


@router.callback_query(F.data.startswith(QM_ANSWER_PREFIX))
async def cb_question_answer(callback: CallbackQuery):
    user_id = callback.from_user.id
    pending = pending_questions.get(user_id)
    if not pending:
        await callback.answer("No pending question.", show_alert=True)
        return

    try:
        idx = int(callback.data[len(QM_ANSWER_PREFIX):])
        answer = pending["options"][idx]
    except (ValueError, IndexError, KeyError):
        await callback.answer("Invalid option.", show_alert=True)
        return

    pending_questions.pop(user_id, None)
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass
    await callback.answer()

    model_key = get_group_model(callback.message.chat.id)
    await send_ai_response(callback.message, answer, model_key, user_id)


# ─── Instruction Guide Command ───────────────────────────────────────────────


@router.message(Command("instruction"))
async def cmd_instruction(message: Message):
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Close", callback_data=INSTR_CLOSE)]
    ])
    await message.answer(INSTRUCTION_GUIDE, reply_markup=kb, parse_mode=ParseMode.HTML)


@router.callback_query(F.data == INSTR_CLOSE)
async def cb_instr_close(callback: CallbackQuery):
    try:
        await callback.message.delete()
    except Exception:
        pass
    await callback.answer("Closed.")


# ─── Log Command ───────────────────────────────────────────────────────────────


@router.message(Command("log"))
async def cmd_log(message: Message):
    if not _is_owner(message.from_user.id):
        await message.answer("Only the bot owner can view logs.")
        return

    if not os.path.exists(LOG_FILE):
        await message.answer("No logs have been recorded yet.")
        return

    size = os.path.getsize(LOG_FILE)
    caption = "Bot logs since startup."

    # Telegram documents are capped; truncate very large logs to the tail.
    if size > 20 * 1024 * 1024:
        with open(LOG_FILE, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
        truncated = os.path.join(os.path.dirname(LOG_FILE), "bot_logs_truncated.txt")
        with open(truncated, "w", encoding="utf-8") as f:
            f.writelines(lines[-2000:])
        await message.answer_document(
            FSInputFile(truncated),
            caption=f"{caption} (last 2000 lines; the full log was too large to send)",
        )
        try:
            os.remove(truncated)
        except Exception:
            pass
        return

    await message.answer_document(FSInputFile(LOG_FILE), caption=caption)


# ─── Startup ─────────────────────────────────────────────────────────────────


async def on_startup():
    await bot.delete_webhook(drop_pending_updates=True)
    me = await bot.get_me()
    bot.username = me.username
    
    # Load image channel ID
    load_image_channel_id()
    
    logger.info(f"Bot started: @{me.username} (id={me.id})")

    user_commands = [
        BotCommand(command="start", description="Start the bot"),
        BotCommand(command="q", description="Ask AI a question"),
        BotCommand(command="search", description="Search the web with AI"),
        BotCommand(command="translate", description="Translate text to another language"),
        BotCommand(command="safety", description="Check content safety"),
        BotCommand(command="post", description="Publish to Telegraph"),
        BotCommand(command="clearhistory", description="Clear your conversation history"),
        BotCommand(command="help", description="Show help"),
        BotCommand(command="model", description="Show current AI model"),
        BotCommand(command="agent", description="Manage your AI agents"),
        BotCommand(command="questionmode", description="Toggle clarifying-question mode"),
        BotCommand(command="instruction", description="Show the getting-started guide"),
    ]
    owner_commands = [
        BotCommand(command="start", description="Start the bot"),
        BotCommand(command="q", description="Ask AI a question"),
        BotCommand(command="search", description="Search the web with AI"),
        BotCommand(command="translate", description="Translate text to another language"),
        BotCommand(command="safety", description="Check content safety"),
        BotCommand(command="post", description="Publish to Telegraph"),
        BotCommand(command="authorize", description="Activate bot in this group"),
        BotCommand(command="deauthorize", description="Deactivate bot in this group"),
        BotCommand(command="unauthorize", description="Deactivate bot in this group"),
        BotCommand(command="settings", description="Change AI model"),
        BotCommand(command="model", description="Show current AI model"),
        BotCommand(command="setchannel", description="Set image channel for /post"),
        BotCommand(command="setname", description="Set bot alias name"),
        BotCommand(command="clearhistory", description="Clear your conversation history"),
        BotCommand(command="clearallhistory", description="Clear everyone's history"),
        BotCommand(command="restart", description="Restart the bot"),
        BotCommand(command="help", description="Show help"),
        BotCommand(command="agent", description="Manage your AI agents"),
        BotCommand(command="questionmode", description="Toggle clarifying-question mode"),
        BotCommand(command="instruction", description="Show the getting-started guide"),
        BotCommand(command="log", description="Send bot logs since startup"),
    ]

    await bot.set_my_commands(user_commands, scope=BotCommandScopeAllPrivateChats())
    await bot.set_my_commands(user_commands, scope=BotCommandScopeAllGroupChats())
    await bot.set_my_commands(owner_commands, scope=BotCommandScopeChat(chat_id=OWNER_ID))

    logger.info(f"Owner ID: {db.get('owner_id', OWNER_ID)}")
    logger.info(f"Authorized groups: {len(db['authorized_groups'])}")
    logger.info(f"Image channel: {IMAGE_CHANNEL_ID or 'Not set'}")
    logger.info("Bot is running!")

    marker_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".restart_marker.json")
    if os.path.exists(marker_path):
        try:
            with open(marker_path, "r", encoding="utf-8") as f:
                marker = json.load(f)
            chat_id = marker.get("chat_id")
            notice_id = marker.get("notice_message_id")
            back_text = f"✅ Bot is back online{_EXC}"
            edited = False
            if chat_id and notice_id:
                try:
                    await bot.edit_message_text(
                        text=back_text,
                        chat_id=chat_id,
                        message_id=notice_id,
                        parse_mode=ParseMode.MARKDOWN_V2,
                    )
                    edited = True
                except Exception as e:
                    logger.warning(f"Could not edit restart notice: {e}")
            if not edited and chat_id:
                try:
                    await bot.send_message(
                        chat_id=chat_id,
                        text=back_text,
                        parse_mode=ParseMode.MARKDOWN_V2,
                    )
                except Exception as e:
                    logger.warning(f"Could not send restart-back notice: {e}")
        except Exception as e:
            logger.warning(f"Failed to process restart marker: {e}")
        finally:
            try:
                os.remove(marker_path)
            except Exception:
                pass


async def main():
    dp.startup.register(on_startup)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
